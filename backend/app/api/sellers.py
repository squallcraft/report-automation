from typing import Optional, List

import io

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from app.database import get_db
from app.auth import require_admin, require_admin_or_administracion, require_permission, hash_password
from app.models import Seller, Sucursal
from app.schemas import SellerCreate, SellerUpdate, SellerOut, SucursalCreate, SucursalOut
from app.services.audit import registrar as audit
from app.services.audit import diff_campos

router = APIRouter(prefix="/sellers", tags=["Sellers"])


@router.get("", response_model=List[SellerOut])
def listar_sellers(
    activo: Optional[bool] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    query = db.query(Seller).options(joinedload(Seller.sucursales))
    if activo is not None:
        query = query.filter(Seller.activo == activo)
    if q:
        query = query.filter(Seller.nombre.ilike(f"%{q}%"))
    return query.order_by(Seller.nombre).all()


@router.get("/plantilla/descargar")
def descargar_plantilla_sellers():
    """Genera plantilla Excel para importación masiva de sellers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sellers"

    columnas = ["Nombre", "Aliases (separados por ;)", "Plan Tarifario", "Tarifa (precio base)", "Zona", "Empresa", "Retiro (monto)", "Tarifa retiro driver", "Min paquetes retiro gratis"]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    for i, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = border

    ejemplos = [
        ["MercadoLibre Chile", "Mercado Libre;ML Chile;MeLi", "2800", 2500, "Santiago", "ECOURIER", 1500, 1000, 30],
        ["Ferretería Oviedo", "Oviedo;FERRETERIA OVIEDO", "Oviedo", 2200, "Santiago", "OVIEDO", 0, 0, 0],
        ["Aventura Store", "AVENTURA STORE;Aventura", "V", 2000, "Santiago", "TERCERIZADO", 0, 0, 0],
    ]

    alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    for ri, fila in enumerate(ejemplos, 2):
        for ci, val in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=10)
            c.border = border
            if ri % 2 == 0:
                c.fill = alt

    anchos = [28, 35, 18, 18, 16, 16, 18, 20, 26]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.auto_filter.ref = "A1:I1"
    ws.freeze_panes = "A2"

    wsi = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("Instrucciones para Importación de Sellers", ""),
        ("", ""),
        ("Campo", "Descripción"),
        ("Nombre", "Nombre oficial del seller (obligatorio, único)."),
        ("Aliases", "Nombres alternativos separados por punto y coma (;). Se usan para homologar el reporte del software."),
        ("Plan Tarifario", "Nombre del plan en la matriz de tarifas (Ej: 2800, V, Giorgio). El cobro se calcula por plan + comuna."),
        ("Tarifa (precio base)", "Monto CLP fallback si el plan/comuna no existe en la matriz."),
        ("Zona", "Zona geográfica: Santiago, Valparaíso, etc."),
        ("Empresa", "ECOURIER, TERCERIZADO u OVIEDO. Determina la tarifa que se paga al driver."),
        ("Retiro (monto)", "Monto que se cobra al seller por cada retiro. Si es 0 o vacío, no tiene servicio de retiro."),
        ("Tarifa retiro driver", "Monto que se paga al driver por cada retiro en este seller."),
        ("Min paquetes retiro gratis", "Si el seller tiene más de X paquetes en la semana, el retiro es gratis. 0 = siempre cobra."),
    ]
    for ri, (a, b) in enumerate(instrucciones, 1):
        ca = wsi.cell(row=ri, column=1, value=a)
        cb = wsi.cell(row=ri, column=2, value=b)
        if ri == 1:
            ca.font = Font(bold=True, size=14, color="1A365D")
        elif ri == 3:
            ca.font = Font(bold=True, size=11)
            cb.font = Font(bold=True, size=11)
        else:
            ca.font = Font(bold=True, size=10) if a else Font(size=10)
            cb.font = Font(size=10, color="4A5568")
    wsi.column_dimensions["A"].width = 30
    wsi.column_dimensions["B"].width = 85

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_sellers.xlsx"})


@router.post("/importar")
async def importar_sellers(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """Importa sellers desde un archivo Excel."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")

    content = await file.read()
    df = pd.read_excel(io.BytesIO(content))

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "nombre" in cl and "alias" not in cl:
            col_map[col] = "nombre"
        elif "alias" in cl:
            col_map[col] = "aliases"
        elif "plan" in cl and "tarif" in cl:
            col_map[col] = "plan_tarifario"
        elif ("tarifa" in cl or "precio" in cl) and "retiro" not in cl and "pago" not in cl and "plan" not in cl:
            col_map[col] = "precio_base"
        elif "zona" in cl:
            col_map[col] = "zona"
        elif "empresa" in cl:
            col_map[col] = "empresa"
        elif "retiro" in cl and ("driver" in cl or "pago" in cl):
            col_map[col] = "tarifa_retiro_driver"
        elif "retiro" in cl and ("monto" in cl or "tarifa" in cl or col_map.get(col) is None):
            if "min" in cl or "paquete" in cl or "gratis" in cl:
                col_map[col] = "min_paquetes"
            else:
                col_map[col] = "tarifa_retiro"
        elif "min" in cl and ("paquete" in cl or "retiro" in cl or "gratis" in cl):
            col_map[col] = "min_paquetes"
        elif "retiro" in cl and "driver" in cl:
            col_map[col] = "tarifa_retiro_driver"
    df = df.rename(columns=col_map)

    creados = 0
    actualizados = 0
    errores = []
    seen_names = set()

    for idx, row in df.iterrows():
        try:
            nombre = str(row.get("nombre", "")).strip()
            if not nombre or nombre == "nan":
                continue
            if nombre in seen_names:
                continue
            seen_names.add(nombre)

            aliases_raw = str(row.get("aliases", "")) if not pd.isna(row.get("aliases", "")) else ""
            aliases = [a.strip() for a in aliases_raw.split(";") if a.strip()] if aliases_raw else []

            precio_base = int(row.get("precio_base", 0)) if not pd.isna(row.get("precio_base", 0)) else 0
            plan_tarifario = str(row.get("plan_tarifario", "")).strip() if not pd.isna(row.get("plan_tarifario", "")) else None
            if plan_tarifario == "" or plan_tarifario == "nan":
                plan_tarifario = None
            zona = str(row.get("zona", "Santiago")).strip() if not pd.isna(row.get("zona", "")) else "Santiago"
            empresa_raw = str(row.get("empresa", "ECOURIER")).strip().upper() if not pd.isna(row.get("empresa", "")) else "ECOURIER"
            if empresa_raw not in ("ECOURIER", "TERCERIZADO", "OVIEDO"):
                empresa_raw = "ECOURIER"

            tarifa_retiro = int(row.get("tarifa_retiro", 0)) if not pd.isna(row.get("tarifa_retiro", 0)) else 0
            tarifa_retiro_driver = int(row.get("tarifa_retiro_driver", 0)) if not pd.isna(row.get("tarifa_retiro_driver", 0)) else 0
            min_paq = int(row.get("min_paquetes", 0)) if not pd.isna(row.get("min_paquetes", 0)) else 0
            tiene_retiro = tarifa_retiro > 0

            existing = db.query(Seller).filter(Seller.nombre == nombre).first()
            if existing:
                existing.aliases = aliases if aliases else existing.aliases
                existing.precio_base = precio_base or existing.precio_base
                if plan_tarifario:
                    existing.plan_tarifario = plan_tarifario
                existing.zona = zona
                existing.empresa = empresa_raw
                existing.tarifa_retiro = tarifa_retiro
                existing.tarifa_retiro_driver = tarifa_retiro_driver
                existing.min_paquetes_retiro_gratis = min_paq
                existing.tiene_retiro = tiene_retiro
                actualizados += 1
            else:
                seller = Seller(
                    nombre=nombre, aliases=aliases, precio_base=precio_base,
                    plan_tarifario=plan_tarifario,
                    zona=zona, empresa=empresa_raw, tarifa_retiro=tarifa_retiro,
                    tarifa_retiro_driver=tarifa_retiro_driver,
                    min_paquetes_retiro_gratis=min_paq, tiene_retiro=tiene_retiro,
                )
                db.add(seller)
                creados += 1
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {str(e)}")

    db.commit()

    audit(db, "importar_sellers", usuario=current_user, request=request,
          entidad="seller",
          metadata={"archivo": file.filename, "creados": creados, "actualizados": actualizados})

    return {"creados": creados, "actualizados": actualizados, "errores": errores}


@router.get("/plantilla/rut-giro/descargar")
def descargar_plantilla_rut_giro(db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    """Plantilla Excel con sellers homologados para completar RUT y Giro."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RUT y Giro"

    columnas = ["Vendedor", "RUT", "Giro"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    for i, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = border

    sellers = db.query(Seller).filter(Seller.activo == True).order_by(Seller.nombre).all()
    alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    for ri, s in enumerate(sellers, 2):
        ws.cell(row=ri, column=1, value=s.nombre).font = Font(size=10, bold=True)
        ws.cell(row=ri, column=2, value=s.rut or "").font = Font(size=10)
        ws.cell(row=ri, column=3, value=s.giro or "").font = Font(size=10)
        for ci in range(1, 4):
            ws.cell(row=ri, column=ci).border = border
            if ri % 2 == 0:
                ws.cell(row=ri, column=ci).fill = alt

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.auto_filter.ref = "A1:C1"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_rut_giro_sellers.xlsx"})


@router.post("/importar/rut-giro")
async def importar_rut_giro(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_administracion),
):
    """Importa RUT y Giro de sellers desde Excel."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")

    content = await file.read()
    df = pd.read_excel(io.BytesIO(content))

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "vendedor" in cl or "seller" in cl or "nombre" in cl:
            col_map[col] = "nombre"
        elif "rut" in cl:
            col_map[col] = "rut"
        elif "giro" in cl:
            col_map[col] = "giro"
    df = df.rename(columns=col_map)

    actualizados = 0
    errores = []

    for idx, row in df.iterrows():
        try:
            nombre = str(row.get("nombre", "")).strip()
            if not nombre or nombre == "nan":
                continue

            seller = db.query(Seller).filter(Seller.nombre == nombre).first()
            if not seller:
                errores.append(f"Fila {idx + 2}: seller '{nombre}' no encontrado")
                continue

            rut = str(row.get("rut", "")).strip() if not pd.isna(row.get("rut", "")) else None
            giro = str(row.get("giro", "")).strip() if not pd.isna(row.get("giro", "")) else None

            if rut and rut != "nan":
                seller.rut = rut
            if giro and giro != "nan":
                seller.giro = giro

            actualizados += 1
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {str(e)}")

    db.commit()

    audit(db, "importar_rut_giro", usuario=current_user, request=request,
          entidad="seller",
          metadata={"archivo": file.filename, "actualizados": actualizados})

    return {"actualizados": actualizados, "errores": errores}


@router.get("/{seller_id}", response_model=SellerOut)
def obtener_seller(seller_id: int, db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    seller = db.query(Seller).options(joinedload(Seller.sucursales)).filter(Seller.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")
    return seller


@router.post("", response_model=SellerOut, status_code=201)
def crear_seller(data: SellerCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(require_permission("sellers:editar"))):
    existing = db.query(Seller).filter(Seller.nombre == data.nombre).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un seller con ese nombre")
    dump = data.model_dump(exclude={"password"})
    if dump.get("email") == "":
        dump["email"] = None
    seller = Seller(**dump)
    if data.password:
        seller.password_hash = hash_password(data.password)
    db.add(seller)
    db.commit()
    db.refresh(seller)

    audit(db, "crear_seller", usuario=current_user, request=request,
          entidad="seller", entidad_id=seller.id,
          metadata={"nombre": seller.nombre, "precio_base": seller.precio_base,
                    "plan_tarifario": seller.plan_tarifario, "empresa": seller.empresa})

    return seller


@router.put("/{seller_id}", response_model=SellerOut)
def actualizar_seller(
    seller_id: int, data: SellerUpdate, request: Request,
    db: Session = Depends(get_db), current_user=Depends(require_permission("sellers:editar")),
):
    seller = db.query(Seller).get(seller_id)
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")

    campos_audit = [
        "nombre", "precio_base", "tarifa_retiro", "aliases", "plan_tarifario",
        "usa_pickup", "tiene_retiro", "rut", "giro", "zona", "empresa",
        "tarifa_retiro_driver", "min_paquetes_retiro_gratis", "email",
        "dir_fiscal", "cmna_fiscal", "correo_dte",
    ]
    old_values = {c: getattr(seller, c, None) for c in campos_audit}

    update_data = data.model_dump(exclude_unset=True, exclude={"password"})
    if update_data.get("email") == "":
        update_data["email"] = None
    for key, value in update_data.items():
        setattr(seller, key, value)
    if data.password:
        seller.password_hash = hash_password(data.password)
    db.flush()

    new_values = {c: getattr(seller, c, None) for c in campos_audit}
    cambios = diff_campos(old_values, new_values, campos_audit)
    if cambios:
        audit(db, "editar_seller", usuario=current_user, request=request,
              entidad="seller", entidad_id=seller.id, cambios=cambios,
              metadata={"nombre": seller.nombre})

    db.commit()
    db.refresh(seller)
    return seller


@router.delete("/{seller_id}")
def eliminar_seller(seller_id: int, request: Request, db: Session = Depends(get_db), current_user=Depends(require_permission("sellers:editar"))):
    seller = db.query(Seller).get(seller_id)
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")
    seller.activo = False
    db.commit()

    audit(db, "eliminar_seller", usuario=current_user, request=request,
          entidad="seller", entidad_id=seller.id,
          metadata={"nombre": seller.nombre})

    return {"message": "Seller desactivado"}


# ── Sucursales ──

@router.get("/{seller_id}/sucursales", response_model=List[SucursalOut])
def listar_sucursales(seller_id: int, db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    seller = db.query(Seller).get(seller_id)
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")
    return db.query(Sucursal).filter(Sucursal.seller_id == seller_id).order_by(Sucursal.nombre).all()


@router.post("/{seller_id}/sucursales", response_model=SucursalOut, status_code=201)
def crear_sucursal(
    seller_id: int, data: SucursalCreate, request: Request,
    db: Session = Depends(get_db), current_user=Depends(require_permission("sellers:editar")),
):
    seller = db.query(Seller).get(seller_id)
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")
    suc = Sucursal(seller_id=seller_id, **data.model_dump())
    db.add(suc)
    db.commit()
    db.refresh(suc)

    audit(db, "crear_sucursal", usuario=current_user, request=request,
          entidad="sucursal", entidad_id=suc.id,
          metadata={"seller_id": seller_id, "nombre": suc.nombre})

    return suc


@router.put("/{seller_id}/sucursales/{suc_id}", response_model=SucursalOut)
def actualizar_sucursal(
    seller_id: int, suc_id: int, data: SucursalCreate, request: Request,
    db: Session = Depends(get_db), current_user=Depends(require_permission("sellers:editar")),
):
    suc = db.query(Sucursal).filter(Sucursal.id == suc_id, Sucursal.seller_id == seller_id).first()
    if not suc:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    for key, value in data.model_dump().items():
        setattr(suc, key, value)
    db.commit()
    db.refresh(suc)

    audit(db, "editar_sucursal", usuario=current_user, request=request,
          entidad="sucursal", entidad_id=suc.id,
          metadata={"seller_id": seller_id, "nombre": suc.nombre})

    return suc


@router.delete("/{seller_id}/sucursales/{suc_id}")
def eliminar_sucursal(
    seller_id: int, suc_id: int, request: Request,
    db: Session = Depends(get_db), current_user=Depends(require_permission("sellers:editar")),
):
    suc = db.query(Sucursal).filter(Sucursal.id == suc_id, Sucursal.seller_id == seller_id).first()
    if not suc:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    db.delete(suc)
    db.commit()

    audit(db, "eliminar_sucursal", usuario=current_user, request=request,
          entidad="sucursal", entidad_id=suc_id,
          metadata={"seller_id": seller_id, "nombre": suc.nombre})

    return {"message": "Sucursal eliminada"}


# ──────────────────────────────────────────────────────────────────────────────
# Lifecycle comercial: cerrar / pausar / reabrir
# ──────────────────────────────────────────────────────────────────────────────
from pydantic import BaseModel as _PB
from datetime import date as _date
from app.models import GestionComercialEntry


class CerrarSellerBody(_PB):
    razones_cierre: List[str] = []
    conversacion_salida: Optional[str] = None     # 'si' | 'no' | 'parcial'
    destino_competencia: Optional[str] = None
    potencial_recuperacion: Optional[str] = None  # 'alto' | 'medio' | 'bajo' | 'ninguno'
    condicion_recuperacion: Optional[str] = None
    nota_cierre: Optional[str] = None


class PausarSellerBody(_PB):
    fecha_pausa_fin: Optional[str] = None   # ISO date
    nota: Optional[str] = None


class ReabrirSellerBody(_PB):
    como_volvio: Optional[str] = None  # 'espontaneo' | 'outreach' | 'oferta'
    que_cambio: Optional[str] = None
    nota: Optional[str] = None


@router.post("/{seller_id}/cerrar")
def cerrar_seller(
    seller_id: int, body: CerrarSellerBody, request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(404, "Seller no encontrado")

    seller.activo = False
    seller.tipo_cierre = "cerrado"
    seller.fecha_cierre = _date.today()
    seller.razones_cierre = body.razones_cierre
    seller.conversacion_salida = body.conversacion_salida
    seller.destino_competencia = body.destino_competencia
    seller.potencial_recuperacion = body.potencial_recuperacion
    seller.condicion_recuperacion = body.condicion_recuperacion
    seller.nota_cierre = body.nota_cierre
    seller.cerrado_por = getattr(current_user, "nombre", None) or getattr(current_user, "email", None)

    nota_log = f"Cliente cerrado. Razones: {', '.join(body.razones_cierre) or '—'}."
    if body.nota_cierre:
        nota_log += f" {body.nota_cierre}"
    db.add(GestionComercialEntry(
        seller_id=seller_id,
        fecha=_date.today(),
        tipo="interno",
        estado="perdido",
        razon=body.razones_cierre[0] if body.razones_cierre else None,
        nota=nota_log,
        usuario=seller.cerrado_por,
    ))
    db.commit()
    audit(db, "cerrar_seller", usuario=current_user, request=request,
          entidad="seller", entidad_id=seller_id,
          metadata={"razones": body.razones_cierre, "potencial": body.potencial_recuperacion})
    return {"ok": True}


@router.post("/{seller_id}/pausar")
def pausar_seller(
    seller_id: int, body: PausarSellerBody, request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(404, "Seller no encontrado")

    seller.activo = False
    seller.tipo_cierre = "pausa"
    seller.fecha_cierre = _date.today()
    seller.fecha_pausa_fin = _date.fromisoformat(body.fecha_pausa_fin) if body.fecha_pausa_fin else None
    seller.nota_cierre = body.nota
    seller.cerrado_por = getattr(current_user, "nombre", None) or getattr(current_user, "email", None)

    db.add(GestionComercialEntry(
        seller_id=seller_id,
        fecha=_date.today(),
        tipo="interno",
        estado="en_pausa",
        nota=f"Seller pausado. {body.nota or ''}".strip(),
        recordatorio=seller.fecha_pausa_fin,
        usuario=seller.cerrado_por,
    ))
    db.commit()
    audit(db, "pausar_seller", usuario=current_user, request=request,
          entidad="seller", entidad_id=seller_id, metadata={})
    return {"ok": True}


@router.post("/{seller_id}/reabrir")
def reabrir_seller(
    seller_id: int, body: ReabrirSellerBody, request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(404, "Seller no encontrado")

    prev_tipo = seller.tipo_cierre
    seller.activo = True
    seller.tipo_cierre = None
    seller.fecha_cierre = None
    seller.fecha_pausa_fin = None
    seller.razones_cierre = []
    seller.conversacion_salida = None
    seller.destino_competencia = None
    seller.potencial_recuperacion = None
    seller.condicion_recuperacion = None
    seller.nota_cierre = None
    seller.cerrado_por = None

    nota_log = f"Seller reactivado (era: {prev_tipo or 'inactivo'})."
    if body.como_volvio:
        nota_log += f" Regreso: {body.como_volvio}."
    if body.que_cambio:
        nota_log += f" Cambio: {body.que_cambio}."
    if body.nota:
        nota_log += f" {body.nota}"
    db.add(GestionComercialEntry(
        seller_id=seller_id,
        fecha=_date.today(),
        tipo="interno",
        estado="recuperado",
        nota=nota_log,
        usuario=getattr(current_user, "nombre", None) or getattr(current_user, "email", None),
    ))
    db.commit()
    audit(db, "reabrir_seller", usuario=current_user, request=request,
          entidad="seller", entidad_id=seller_id, metadata={"como_volvio": body.como_volvio})
    return {"ok": True}


@router.get("/no-activos")
def listar_no_activos(
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    """Retorna sellers pausados y cerrados con su metadata de cierre."""
    sellers = db.query(Seller).filter(
        Seller.activo == False,
        Seller.tipo_cierre.isnot(None),
    ).order_by(Seller.fecha_cierre.desc()).all()

    pausados = []
    cerrados = []
    for s in sellers:
        row = {
            "id": s.id,
            "nombre": s.nombre,
            "empresa": s.empresa,
            "rut": s.rut,
            "tipo_cierre": s.tipo_cierre,
            "fecha_cierre": s.fecha_cierre.isoformat() if s.fecha_cierre else None,
            "fecha_pausa_fin": s.fecha_pausa_fin.isoformat() if s.fecha_pausa_fin else None,
            "razones_cierre": s.razones_cierre or [],
            "conversacion_salida": s.conversacion_salida,
            "destino_competencia": s.destino_competencia,
            "potencial_recuperacion": s.potencial_recuperacion,
            "condicion_recuperacion": s.condicion_recuperacion,
            "nota_cierre": s.nota_cierre,
            "cerrado_por": s.cerrado_por,
        }
        if s.tipo_cierre == "pausa":
            pausados.append(row)
        else:
            cerrados.append(row)

    return {"pausados": pausados, "cerrados": cerrados}


@router.get("/churn-analytics")
def churn_analytics(
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    """Estadísticas de churn: razones, potencial recuperación, tasa de recovery."""
    cerrados = db.query(Seller).filter(
        Seller.activo == False, Seller.tipo_cierre == "cerrado"
    ).all()

    from collections import Counter
    todas_razones = []
    potenciales = Counter()
    for s in cerrados:
        todas_razones.extend(s.razones_cierre or [])
        if s.potencial_recuperacion:
            potenciales[s.potencial_recuperacion] += 1

    # Recovery rate: sellers that were cerrado but are now activo (reopened)
    # We detect this by gestion entries with estado='recuperado' for sellers that are now activo
    total_cerrados_historico = len(cerrados)
    recuperados = db.query(GestionComercialEntry).filter(
        GestionComercialEntry.estado == "recuperado",
    ).count()

    return {
        "total_cerrados": len(cerrados),
        "razones": [{"razon": r, "count": c} for r, c in Counter(todas_razones).most_common()],
        "potencial_recuperacion": dict(potenciales),
        "recuperables": sum(1 for s in cerrados if s.potencial_recuperacion in ("alto", "medio")),
    }

