from typing import Optional, List

from datetime import datetime, timezone, date
from collections import defaultdict
import io
import zipfile

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app.database import get_db
from app.auth import (
    require_admin,
    require_admin_or_administracion,
    get_current_user,
    driver_period_allowed,
    DRIVER_CUTOFF_ANIO,
    DRIVER_CUTOFF_MES,
    DRIVER_CUTOFF_SEMANA,
)
from app.models import (
    PeriodoLiquidacion, EstadoLiquidacionEnum,
    Envio, Seller, Driver, Retiro, AjusteLiquidacion,
    TipoEntidadEnum, EmpresaEnum, ProductoConExtra, TarifaPlanComuna,
    TarifaComuna,
)
from app.services.calendario import get_dates_for_week
from app.services.audit import registrar as audit
from app.schemas import (
    LiquidacionSellerOut, LiquidacionDriverOut, RentabilidadSellerOut,
    PeriodoOut, PeriodoUpdate,
)
from app.services.liquidacion import (
    calcular_liquidacion_sellers, calcular_liquidacion_drivers, calcular_rentabilidad,
)
from app.services.pdf_generator import generar_pdf_seller, generar_pdf_driver

router = APIRouter(prefix="/liquidacion", tags=["Liquidación"])


def _get_or_create_periodo(db: Session, semana: int, mes: int, anio: int) -> PeriodoLiquidacion:
    periodo = db.query(PeriodoLiquidacion).filter(
        PeriodoLiquidacion.semana == semana,
        PeriodoLiquidacion.mes == mes,
        PeriodoLiquidacion.anio == anio,
    ).first()
    if not periodo:
        periodo = PeriodoLiquidacion(semana=semana, mes=mes, anio=anio)
        db.add(periodo)
        db.flush()
    return periodo


@router.get("/sellers", response_model=List[LiquidacionSellerOut])
def liquidacion_sellers(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    periodo = _get_or_create_periodo(db, semana, mes, anio)
    if periodo.snapshot_sellers:
        return periodo.snapshot_sellers

    resultado = calcular_liquidacion_sellers(db, semana, mes, anio)
    periodo.snapshot_sellers = resultado
    flag_modified(periodo, "snapshot_sellers")
    db.commit()
    return resultado


@router.get("/drivers", response_model=List[LiquidacionDriverOut])
def liquidacion_drivers(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    periodo = _get_or_create_periodo(db, semana, mes, anio)
    if periodo.snapshot_drivers:
        return periodo.snapshot_drivers

    resultado = calcular_liquidacion_drivers(db, semana, mes, anio)
    periodo.snapshot_drivers = resultado
    flag_modified(periodo, "snapshot_drivers")
    db.commit()
    return resultado


@router.get("/rentabilidad", response_model=List[RentabilidadSellerOut])
def rentabilidad(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    periodo = _get_or_create_periodo(db, semana, mes, anio)
    if periodo.snapshot_rentabilidad:
        return periodo.snapshot_rentabilidad

    resultado = calcular_rentabilidad(db, semana, mes, anio)
    periodo.snapshot_rentabilidad = resultado
    flag_modified(periodo, "snapshot_rentabilidad")
    db.commit()
    return resultado


@router.post("/recalcular")
def recalcular_liquidacion(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """Fuerza recálculo de la liquidación: reaplicar tarifas actuales y recalcular snapshots."""
    # Reaplicar tarifas actuales de sellers y drivers a los envíos del período
    envios_actualizados = _reaplicar_tarifas(db, semana, mes, anio)

    periodo = _get_or_create_periodo(db, semana, mes, anio)

    periodo.snapshot_sellers = calcular_liquidacion_sellers(db, semana, mes, anio)
    periodo.snapshot_drivers = calcular_liquidacion_drivers(db, semana, mes, anio)
    periodo.snapshot_rentabilidad = calcular_rentabilidad(db, semana, mes, anio)
    flag_modified(periodo, "snapshot_sellers")
    flag_modified(periodo, "snapshot_drivers")
    flag_modified(periodo, "snapshot_rentabilidad")
    db.commit()

    audit(
        db, "recalcular_liquidacion",
        usuario=current_user, request=request,
        entidad="periodo_liquidacion", entidad_id=periodo.id,
        metadata={
            "semana": semana, "mes": mes, "anio": anio,
            "sellers": len(periodo.snapshot_sellers),
            "drivers": len(periodo.snapshot_drivers),
            "envios_tarifas_actualizadas": envios_actualizados,
        },
    )

    return {
        "message": "Liquidación recalculada",
        "sellers": len(periodo.snapshot_sellers),
        "drivers": len(periodo.snapshot_drivers),
        "envios_tarifas_actualizadas": envios_actualizados,
    }


def _reaplicar_tarifas(db: Session, semana: int, mes: int, anio: int) -> int:
    """Reaplicar tarifas actuales de sellers, drivers, comunas y productos a envíos pendientes."""
    envios = db.query(Envio).filter(
        Envio.semana == semana, Envio.mes == mes, Envio.anio == anio,
        Envio.estado_financiero == "pendiente",
    ).all()

    if not envios:
        return 0

    seller_ids = {e.seller_id for e in envios if e.seller_id}
    driver_ids = {e.driver_id for e in envios if e.driver_id}
    sellers_map = {s.id: s for s in db.query(Seller).filter(Seller.id.in_(seller_ids)).all()}
    drivers_map = {d.id: d for d in db.query(Driver).filter(Driver.id.in_(driver_ids)).all()}

    tarifas_plan = db.query(TarifaPlanComuna).all()
    plan_comuna_map = {(t.plan_tarifario.lower(), t.comuna.lower()): t.precio for t in tarifas_plan}

    comunas_extra = db.query(TarifaComuna).all()
    comunas_map = {t.comuna.lower(): t for t in comunas_extra}

    productos_extra = db.query(ProductoConExtra).filter(ProductoConExtra.activo == True).all()
    productos_map = {p.codigo_mlc.upper(): p for p in productos_extra}

    actualizados = 0
    for envio in envios:
        cambio = False

        # Tarifa base seller
        seller = sellers_map.get(envio.seller_id)
        if seller:
            if seller.plan_tarifario and envio.comuna:
                key = (seller.plan_tarifario.lower(), envio.comuna.lower())
                nuevo_cobro = plan_comuna_map.get(key, seller.precio_base)
            else:
                nuevo_cobro = seller.precio_base
            if envio.cobro_seller != nuevo_cobro:
                envio.cobro_seller = nuevo_cobro
                cambio = True

        # Tarifa base driver
        driver = drivers_map.get(envio.driver_id)
        if driver and envio.empresa:
            emp = envio.empresa
            if emp in (EmpresaEnum.ECOURIER, EmpresaEnum.ECOURIER.value):
                nuevo_costo = driver.tarifa_ecourier
            elif emp in (EmpresaEnum.OVIEDO, EmpresaEnum.OVIEDO.value):
                nuevo_costo = driver.tarifa_oviedo
            elif emp in (EmpresaEnum.TERCERIZADO, EmpresaEnum.TERCERIZADO.value):
                nuevo_costo = driver.tarifa_tercerizado
            else:
                nuevo_costo = None
            if nuevo_costo is not None and envio.costo_driver != nuevo_costo:
                envio.costo_driver = nuevo_costo
                cambio = True

        # Extra por comuna
        new_extra_comuna_seller = 0
        new_extra_comuna_driver = 0
        if envio.comuna:
            tc = comunas_map.get(envio.comuna.lower())
            if tc:
                new_extra_comuna_seller = tc.extra_seller
                new_extra_comuna_driver = tc.extra_driver
        if envio.extra_comuna_seller != new_extra_comuna_seller:
            envio.extra_comuna_seller = new_extra_comuna_seller
            cambio = True
        if envio.extra_comuna_driver != new_extra_comuna_driver:
            envio.extra_comuna_driver = new_extra_comuna_driver
            cambio = True

        # Extra por producto
        new_extra_prod_seller = 0
        new_extra_prod_driver = 0
        if envio.codigo_producto:
            prod = productos_map.get(envio.codigo_producto.upper())
            if prod:
                new_extra_prod_seller = prod.extra_seller
                new_extra_prod_driver = prod.extra_driver
        if envio.extra_producto_seller != new_extra_prod_seller:
            envio.extra_producto_seller = new_extra_prod_seller
            cambio = True
        if envio.extra_producto_driver != new_extra_prod_driver:
            envio.extra_producto_driver = new_extra_prod_driver
            cambio = True

        # Regla: si driver tiene extra por producto Y por comuna, solo el mayor
        if envio.extra_producto_driver > 0 and envio.extra_comuna_driver > 0:
            if envio.extra_producto_driver >= envio.extra_comuna_driver:
                envio.extra_comuna_driver = 0
            else:
                envio.extra_producto_driver = 0

        # Conductores contratados no reciben extras
        if driver and getattr(driver, 'contratado', False):
            envio.extra_producto_driver = 0
            envio.extra_comuna_driver = 0

        if cambio:
            actualizados += 1

    if actualizados:
        db.flush()

    return actualizados


@router.get("/pdf/seller/{seller_id}")
def descargar_pdf_seller(
    seller_id: int,
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    try:
        pdf_bytes = generar_pdf_seller(db, seller_id, semana, mes, anio)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=liquidacion_seller_{seller_id}_S{semana}.pdf"},
    )


@router.get("/pdf/driver/{driver_id}")
def descargar_pdf_driver(
    driver_id: int,
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    if not driver_period_allowed(anio, mes, semana):
        raise HTTPException(
            status_code=403,
            detail="Para drivers solo se muestra información desde la semana 4 de febrero 2026.",
        )
    try:
        pdf_bytes = generar_pdf_driver(db, driver_id, semana, mes, anio)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=liquidacion_driver_{driver_id}_S{semana}.pdf"},
    )


# ── Detalle ──

def _seller_detail(db: Session, seller_id: int, mes: int, anio: int):
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")

    weekly = {}
    for s in range(1, 6):
        envios = db.query(Envio).filter(
            Envio.seller_id == seller_id, Envio.semana == s,
            Envio.mes == mes, Envio.anio == anio,
        ).all()

        ajustes = db.query(AjusteLiquidacion).filter(
            AjusteLiquidacion.tipo == TipoEntidadEnum.SELLER,
            AjusteLiquidacion.entidad_id == seller_id,
            AjusteLiquidacion.semana == s,
            AjusteLiquidacion.mes == mes,
            AjusteLiquidacion.anio == anio,
        ).all()

        total_retiros = 0
        if seller.tiene_retiro and not seller.usa_pickup and envios and seller.tarifa_retiro:
            if not (seller.min_paquetes_retiro_gratis > 0 and len(envios) >= seller.min_paquetes_retiro_gratis):
                dias_con_envios = len({e.fecha_entrega for e in envios if e.fecha_entrega})
                total_retiros = seller.tarifa_retiro * dias_con_envios

        weekly[s] = {
            "monto": sum(e.cobro_seller for e in envios),
            "envios": len(envios),
            "bultos_extra": sum(e.extra_producto_seller for e in envios),
            "cobro_extra_manual": sum(e.cobro_extra_manual for e in envios),
            "retiros": total_retiros,
            "peso_extra": sum(e.extra_comuna_seller for e in envios),
            "ajustes": sum(a.monto for a in ajustes),
        }

    return {
        "nombre": seller.nombre,
        "empresa": seller.empresa or "",
        "weekly": weekly,
    }


def _driver_detail(db: Session, driver_id: int, mes: int, anio: int):
    driver = db.get(Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")

    weekly = {}
    for s in range(1, 6):
        envios = db.query(Envio).filter(
            Envio.driver_id == driver_id, Envio.semana == s,
            Envio.mes == mes, Envio.anio == anio,
        ).all()
        retiros_q = db.query(Retiro).filter(
            Retiro.driver_id == driver_id, Retiro.semana == s,
            Retiro.mes == mes, Retiro.anio == anio,
        ).all()
        normal = [e for e in envios if e.empresa in (None, "", EmpresaEnum.ECOURIER, EmpresaEnum.ECOURIER.value)]
        oviedo = [e for e in envios if e.empresa in (EmpresaEnum.OVIEDO, EmpresaEnum.OVIEDO.value)]
        tercerizado = [e for e in envios if e.empresa in (EmpresaEnum.TERCERIZADO, EmpresaEnum.TERCERIZADO.value)]

        ajustes = db.query(AjusteLiquidacion).filter(
            AjusteLiquidacion.tipo == TipoEntidadEnum.DRIVER,
            AjusteLiquidacion.entidad_id == driver_id,
            AjusteLiquidacion.semana == s,
            AjusteLiquidacion.mes == mes,
            AjusteLiquidacion.anio == anio,
        ).all()
        bonif = sum(a.monto for a in ajustes if a.monto > 0)
        desc = sum(a.monto for a in ajustes if a.monto < 0)

        es_contratado = getattr(driver, 'contratado', False)
        pago_extra_envios = sum(e.pago_extra_manual for e in envios)
        weekly[s] = {
            "normal_count": len(normal),
            "normal_total": sum(e.costo_driver for e in normal),
            "oviedo_count": len(oviedo),
            "oviedo_total": sum(e.costo_driver for e in oviedo),
            "tercerizado_count": len(tercerizado),
            "tercerizado_total": sum(e.costo_driver for e in tercerizado),
            "comuna": 0 if es_contratado else sum(e.extra_comuna_driver for e in envios),
            "bultos_extra": 0 if es_contratado else sum(e.extra_producto_driver for e in envios) + pago_extra_envios,
            "retiros": sum(r.tarifa_driver for r in retiros_q),
            "bonificaciones": bonif,
            "descuentos": desc,
            "envios": len(envios),
        }

    return {
        "nombre": driver.nombre,
        "tarifa_ecourier": driver.tarifa_ecourier,
        "tarifa_oviedo": driver.tarifa_oviedo,
        "tarifa_tercerizado": driver.tarifa_tercerizado,
        "contratado": getattr(driver, 'contratado', False),
        "weekly": weekly,
    }


def _daily_breakdown(envios_list, retiros_list, field_extra, field_comuna, semana, mes, anio, is_seller=True, db=None, contratado=False, seller=None):
    all_dates = get_dates_for_week(db, semana, mes, anio) if db else []

    daily_map = {}
    for d in all_dates:
        daily_map[d] = {
            "fecha": str(d),
            "envios": 0,
            "bultos_extra": 0,
            "retiros": 0,
            "peso_extra": 0,
            "monto": 0,
        }

    for e in envios_list:
        d = e.fecha_entrega
        if d not in daily_map:
            daily_map[d] = {"fecha": str(d), "envios": 0, "bultos_extra": 0, "retiros": 0, "peso_extra": 0, "monto": 0}
        daily_map[d]["envios"] += 1
        if is_seller:
            daily_map[d]["peso_extra"] += getattr(e, field_comuna, 0)
            daily_map[d]["monto"] += e.cobro_seller
            daily_map[d]["bultos_extra"] += getattr(e, field_extra, 0) + (e.cobro_extra_manual or 0)
        else:
            daily_map[d]["monto"] += e.costo_driver
            if not contratado:
                daily_map[d]["bultos_extra"] += getattr(e, field_extra, 0) + (e.pago_extra_manual or 0)
                daily_map[d]["peso_extra"] += getattr(e, field_comuna, 0)

    # Retiros solo aplican al desglose de drivers (tarifa_driver por parada)
    if not is_seller and retiros_list:
        for r in retiros_list:
            if r.fecha in daily_map:
                daily_map[r.fecha]["retiros"] += r.tarifa_driver

    # monto diario = cobro/pago base + retiro del día (solo drivers)
    for info in daily_map.values():
        info["monto"] = info["monto"] + info["retiros"]

    return sorted(daily_map.values(), key=lambda x: x["fecha"])


def _productos_envios(db: Session, envios_list):
    result = []

    codigos = {e.codigo_producto for e in envios_list if e.codigo_producto}
    if codigos:
        productos = {p.codigo_mlc: p for p in db.query(ProductoConExtra).filter(ProductoConExtra.codigo_mlc.in_(codigos)).all()}
        counter = defaultdict(int)
        for e in envios_list:
            if e.codigo_producto:
                counter[e.codigo_producto] += 1
        for cod, cnt in sorted(counter.items(), key=lambda x: -x[1]):
            p = productos.get(cod)
            extra_s = p.extra_seller if p else 0
            extra_d = p.extra_driver if p else 0
            if extra_s > 0 or extra_d > 0:
                result.append({
                    "codigo_mlc": cod,
                    "descripcion": p.descripcion if p else "",
                    "extra_seller": extra_s,
                    "extra_driver": extra_d,
                    "cantidad": cnt,
                })

    for e in envios_list:
        cobro_m = e.cobro_extra_manual or 0
        pago_m = e.pago_extra_manual or 0
        if cobro_m > 0 or pago_m > 0:
            desc_parts = []
            if e.descripcion_producto:
                desc_parts.append(e.descripcion_producto)
            if e.comuna:
                desc_parts.append(e.comuna)
            if e.bultos and e.bultos > 1:
                desc_parts.append(f"{e.bultos} bultos")
            result.append({
                "codigo_mlc": e.tracking_id or f"Envío #{e.id}",
                "descripcion": f"Extra manual — {', '.join(desc_parts)}" if desc_parts else "Extra manual",
                "extra_seller": cobro_m,
                "extra_driver": pago_m,
                "cantidad": 1,
            })

    return result


@router.get("/detalle/seller/{seller_id}")
def detalle_seller(
    seller_id: int,
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    detail = _seller_detail(db, seller_id, mes, anio)
    seller = db.get(Seller, seller_id)
    envios_semana = db.query(Envio).filter(
        Envio.seller_id == seller_id, Envio.semana == semana,
        Envio.mes == mes, Envio.anio == anio,
    ).order_by(Envio.fecha_entrega).all()
    detail["daily"] = _daily_breakdown(
        envios_semana, [],
        "extra_producto_seller", "extra_comuna_seller",
        semana, mes, anio, is_seller=True, db=db,
    )
    detail["productos"] = _productos_envios(db, envios_semana)
    return detail


@router.get("/detalle/driver/{driver_id}")
def detalle_driver(
    driver_id: int,
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    detail = _driver_detail(db, driver_id, mes, anio)
    driver = db.get(Driver, driver_id)
    es_contratado = getattr(driver, 'contratado', False) if driver else False
    envios_semana = db.query(Envio).filter(
        Envio.driver_id == driver_id, Envio.semana == semana,
        Envio.mes == mes, Envio.anio == anio,
    ).order_by(Envio.fecha_entrega).all()
    retiros_semana = db.query(Retiro).filter(
        Retiro.driver_id == driver_id, Retiro.semana == semana,
        Retiro.mes == mes, Retiro.anio == anio,
    ).all()
    detail["daily"] = _daily_breakdown(
        envios_semana, retiros_semana,
        "extra_producto_driver", "extra_comuna_driver",
        semana, mes, anio, is_seller=False, db=db, contratado=es_contratado,
    )
    detail["productos"] = [] if es_contratado else _productos_envios(db, envios_semana)
    return detail


# ── Exportar envíos a Excel ──

@router.get("/exportar/envios")
def exportar_envios(
    seller_id: Optional[int] = None,
    driver_id: Optional[int] = None,
    semana: Optional[int] = None,
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(Envio).filter(Envio.mes == mes, Envio.anio == anio)
    if semana is not None:
        query = query.filter(Envio.semana == semana)
    if seller_id:
        query = query.filter(Envio.seller_id == seller_id)
    if driver_id:
        query = query.filter(Envio.driver_id == driver_id)
        # Para drivers, febrero 2026 solo desde semana 4 (igual que en el front)
        if mes == DRIVER_CUTOFF_MES and anio == DRIVER_CUTOFF_ANIO:
            query = query.filter(Envio.semana >= DRIVER_CUTOFF_SEMANA)
    envios = query.order_by(Envio.fecha_entrega, Envio.semana).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos"

    headers = [
        "Fecha Entrega", "Fecha Carga", "Tracking", "Seller Code", "External ID",
        "Seller", "Driver",
        "Comuna", "Dirección", "Bultos", "Descripción Producto", "Código MLC",
        "Costo Orden", "Cobro Seller", "Extra Prod. Seller", "Extra Com. Seller",
        "Extra Manual Seller", "Pago Driver", "Extra Prod. Driver", "Extra Com. Driver",
        "Extra Manual Driver", "Empresa", "Ruta",
    ]
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        c.border = thin

    for idx, e in enumerate(envios, 2):
        seller = db.get(Seller, e.seller_id) if e.seller_id else None
        driver = db.get(Driver, e.driver_id) if e.driver_id else None
        row = [
            str(e.fecha_entrega) if e.fecha_entrega else "",
            str(e.fecha_carga) if e.fecha_carga else "",
            e.tracking_id or "",
            e.seller_code or "",
            e.venta_id or "",
            seller.nombre if seller else (e.seller_nombre_raw or ""),
            driver.nombre if driver else (e.driver_nombre_raw or ""),
            e.comuna or "",
            e.direccion or "",
            e.bultos,
            e.descripcion_producto or "",
            e.codigo_producto or "",
            e.costo_orden,
            e.cobro_seller,
            e.extra_producto_seller,
            e.extra_comuna_seller,
            e.cobro_extra_manual,
            e.costo_driver,
            e.extra_producto_driver,
            e.extra_comuna_driver,
            e.pago_extra_manual,
            e.empresa or "",
            e.ruta_nombre or "",
        ]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=idx, column=ci, value=v)
            c.border = thin

    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    name = "envios"
    if seller_id:
        s = db.get(Seller, seller_id)
        name = f"envios_{s.nombre}" if s else name
    elif driver_id:
        d = db.get(Driver, driver_id)
        name = f"envios_{d.nombre}" if d else name

    period_suffix = f"_S{semana}_M{mes}_{anio}" if semana is not None else f"_M{mes}_{anio}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={name}{period_suffix}.xlsx"},
    )


# ── ZIP masivo de PDFs ──

@router.get("/zip/sellers")
def descargar_zip_sellers(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    resultado = calcular_liquidacion_sellers(db, semana, mes, anio)
    if not resultado:
        raise HTTPException(status_code=404, detail="No hay datos de sellers para este período")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in resultado:
            try:
                pdf = generar_pdf_seller(db, item["seller_id"], semana, mes, anio)
                nombre = item["seller_nombre"].replace("/", "-").replace("\\", "-")
                zf.writestr(f"{nombre}_S{semana}.pdf", pdf)
            except Exception:
                pass
    size = buf.tell()
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=liquidacion_sellers_S{semana}_M{mes}_{anio}.zip",
            "Content-Length": str(size),
        },
    )


@router.get("/zip/drivers")
def descargar_zip_drivers(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    if not driver_period_allowed(anio, mes, semana):
        raise HTTPException(
            status_code=403,
            detail="Para drivers solo se incluye información desde la semana 4 de febrero 2026.",
        )
    resultado = calcular_liquidacion_drivers(db, semana, mes, anio)
    if not resultado:
        raise HTTPException(status_code=404, detail="No hay datos de drivers para este período")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in resultado:
            try:
                pdf = generar_pdf_driver(db, item["driver_id"], semana, mes, anio)
                nombre = item["driver_nombre"].replace("/", "-").replace("\\", "-")
                zf.writestr(f"{nombre}_S{semana}.pdf", pdf)
            except Exception:
                pass
    size = buf.tell()
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename=liquidacion_drivers_S{semana}_M{mes}_{anio}.zip",
            "Content-Length": str(size),
        },
    )


# ── Períodos ──

@router.get("/periodos", response_model=List[PeriodoOut])
def listar_periodos(
    anio: Optional[int] = None,
    mes: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    query = db.query(PeriodoLiquidacion)
    if anio:
        query = query.filter(PeriodoLiquidacion.anio == anio)
    if mes:
        query = query.filter(PeriodoLiquidacion.mes == mes)
    return query.order_by(PeriodoLiquidacion.anio.desc(), PeriodoLiquidacion.mes.desc(), PeriodoLiquidacion.semana.desc()).all()


@router.post("/periodos", response_model=PeriodoOut, status_code=201)
def crear_periodo(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    existing = db.query(PeriodoLiquidacion).filter(
        PeriodoLiquidacion.semana == semana,
        PeriodoLiquidacion.mes == mes,
        PeriodoLiquidacion.anio == anio,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un período para esa semana")
    periodo = PeriodoLiquidacion(semana=semana, mes=mes, anio=anio)
    db.add(periodo)
    db.commit()
    db.refresh(periodo)
    return periodo


@router.put("/periodos/{periodo_id}", response_model=PeriodoOut)
def actualizar_periodo(
    periodo_id: int,
    data: PeriodoUpdate,
    request: Request,
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    periodo = db.query(PeriodoLiquidacion).get(periodo_id)
    if not periodo:
        raise HTTPException(status_code=404, detail="Período no encontrado")

    estado_anterior = periodo.estado
    periodo.estado = data.estado

    if data.estado == EstadoLiquidacionEnum.APROBADO:
        periodo.aprobado_por = admin["nombre"]
        periodo.aprobado_en = datetime.now(timezone.utc)

    # Fase 3: al aprobar un período, marcar envíos como liquidados
    if data.estado in (EstadoLiquidacionEnum.APROBADO, EstadoLiquidacionEnum.PAGADO):
        envios_afectados = db.query(Envio).filter(
            Envio.semana == periodo.semana,
            Envio.mes == periodo.mes,
            Envio.anio == periodo.anio,
            Envio.is_liquidado == False,
        ).all()
        for e in envios_afectados:
            e.is_liquidado = True
            e.sync_estado_financiero()

        audit(
            db, "cerrar_semana",
            usuario=admin, request=request,
            entidad="periodo_liquidacion", entidad_id=periodo_id,
            cambios={"estado": {"antes": estado_anterior, "despues": data.estado}},
            metadata={
                "semana": periodo.semana, "mes": periodo.mes, "anio": periodo.anio,
                "envios_bloqueados": len(envios_afectados),
            },
        )

    db.commit()
    db.refresh(periodo)
    return periodo


@router.get("/mi-pdf")
def descargar_mi_pdf_driver(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Permite a un driver descargar su propio PDF de liquidación."""
    from app.models import RolEnum
    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")
    if not driver_period_allowed(anio, mes, semana):
        raise HTTPException(
            status_code=403,
            detail="Solo puedes descargar información desde la semana 4 de febrero 2026 en adelante.",
        )
    try:
        pdf_bytes = generar_pdf_driver(db, user["id"], semana, mes, anio)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=mi_liquidacion_S{semana}.pdf"},
    )
