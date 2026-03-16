from typing import List

import io
import os
import shutil

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from app.database import get_db
from app.auth import require_admin, require_admin_or_administracion
from app.models import ProductoConExtra, Envio
from app.schemas import ProductoExtraCreate, ProductoExtraUpdate, ProductoExtraOut

router = APIRouter(prefix="/productos", tags=["Productos con Extra"])


def _recalcular_extras_envios(db: Session, codigos: set[str]):
    """
    Recalcula extra_producto_seller/driver en envíos que tengan un código afectado,
    EXCEPTO aquellos que pertenecen a semanas ya cerradas (PagoSemanaDriver/Seller = PAGADO).
    Aplica regla de negocio: si driver tiene extra producto Y extra comuna,
    se mantiene solo el mayor.
    """
    if not codigos:
        return 0

    from app.models import PagoSemanaDriver, PagoSemanaSeller, EstadoPagoEnum

    # Obtener semanas cerradas para drivers y sellers (inmutables)
    semanas_cerradas_driver: set = set()
    for r in db.query(PagoSemanaDriver).filter(
        PagoSemanaDriver.estado == EstadoPagoEnum.PAGADO.value
    ).all():
        semanas_cerradas_driver.add((r.driver_id, r.semana, r.mes, r.anio))

    semanas_cerradas_seller: set = set()
    for r in db.query(PagoSemanaSeller).filter(
        PagoSemanaSeller.estado == EstadoPagoEnum.PAGADO.value
    ).all():
        semanas_cerradas_seller.add((r.seller_id, r.semana, r.mes, r.anio))

    envios = db.query(Envio).filter(Envio.codigo_producto.in_(codigos)).all()
    updated = 0
    skipped = 0
    for envio in envios:
        # No tocar envíos de semanas cerradas
        driver_cerrado = (envio.driver_id, envio.semana, envio.mes, envio.anio) in semanas_cerradas_driver
        seller_cerrado = (envio.seller_id, envio.semana, envio.mes, envio.anio) in semanas_cerradas_seller
        if driver_cerrado or seller_cerrado:
            skipped += 1
            continue

        prod = db.query(ProductoConExtra).filter(
            ProductoConExtra.codigo_mlc == envio.codigo_producto,
            ProductoConExtra.activo == True,
        ).first()
        new_seller = prod.extra_seller if prod else 0
        new_driver = prod.extra_driver if prod else 0

        # Regla: si driver tiene ambos extras, aplicar solo el mayor
        extra_comuna_d = envio.extra_comuna_driver or 0
        if new_driver > 0 and extra_comuna_d > 0:
            if new_driver >= extra_comuna_d:
                envio.extra_comuna_driver = 0
            else:
                new_driver = 0

        if envio.extra_producto_seller != new_seller or envio.extra_producto_driver != new_driver:
            envio.extra_producto_seller = new_seller
            envio.extra_producto_driver = new_driver
            updated += 1
    return updated


@router.get("/plantilla")
def descargar_plantilla_productos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Extra"

    columnas = ["Código MLC", "Descripción", "Extra Seller (CLP)", "Extra Driver (CLP)"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col_idx, nombre in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ejemplo = [
        ["MLC1774402962", "Producto voluminoso tipo A", 2200, 1000],
        ["MLC2220238846", "Crema Reductora especial", 2200, 1000],
        ["MLC1850003921", "Electrodoméstico grande", 3000, 1500],
    ]
    data_font = Font(size=10)
    data_align = Alignment(vertical="center")
    alt_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

    for row_idx, fila in enumerate(ejemplo, 2):
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    anchos = [22, 35, 20, 20]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho

    ws.auto_filter.ref = "A1:D1"
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos_extra.xlsx"},
    )


@router.post("/importar")
async def importar_productos(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")

    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "código" in cl or "codigo" in cl:
            col_map[col] = "codigo_mlc"
        elif "descripci" in cl:
            col_map[col] = "descripcion"
        elif "seller" in cl:
            col_map[col] = "extra_seller"
        elif "driver" in cl:
            col_map[col] = "extra_driver"
    df = df.rename(columns=col_map)

    creados = 0
    actualizados = 0
    errores = []
    seen = set()

    for idx, row in df.iterrows():
        try:
            codigo = str(row.get("codigo_mlc", "")).strip()
            if not codigo or codigo == "nan":
                continue
            if codigo in seen:
                continue
            seen.add(codigo)

            descripcion = str(row.get("descripcion", "")).strip() if not pd.isna(row.get("descripcion")) else ""
            extra_seller = int(row.get("extra_seller", 0)) if not pd.isna(row.get("extra_seller")) else 0
            extra_driver = int(row.get("extra_driver", 0)) if not pd.isna(row.get("extra_driver")) else 0

            existing = db.query(ProductoConExtra).filter(ProductoConExtra.codigo_mlc == codigo).first()
            if existing:
                existing.descripcion = descripcion or existing.descripcion
                existing.extra_seller = extra_seller
                existing.extra_driver = extra_driver
                existing.activo = True
                actualizados += 1
            else:
                db.add(ProductoConExtra(
                    codigo_mlc=codigo,
                    descripcion=descripcion,
                    extra_seller=extra_seller,
                    extra_driver=extra_driver,
                ))
                db.flush()
                creados += 1
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {e}")

    db.commit()

    envios_actualizados = _recalcular_extras_envios(db, seen)
    db.commit()

    return {
        "creados": creados,
        "actualizados": actualizados,
        "envios_recalculados": envios_actualizados,
        "errores": errores,
    }


@router.get("", response_model=List[ProductoExtraOut])
def listar_productos(db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    return db.query(ProductoConExtra).order_by(ProductoConExtra.codigo_mlc).all()


@router.post("", response_model=ProductoExtraOut, status_code=201)
def crear_producto(data: ProductoExtraCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    existing = db.query(ProductoConExtra).filter(ProductoConExtra.codigo_mlc == data.codigo_mlc).first()
    if existing:
        existing.descripcion = data.descripcion or existing.descripcion
        existing.extra_seller = data.extra_seller
        existing.extra_driver = data.extra_driver
        existing.activo = True
        db.commit()
        _recalcular_extras_envios(db, {existing.codigo_mlc})
        db.commit()
        db.refresh(existing)
        return existing
    producto = ProductoConExtra(**data.model_dump())
    db.add(producto)
    db.commit()
    _recalcular_extras_envios(db, {producto.codigo_mlc})
    db.commit()
    db.refresh(producto)
    return producto


@router.put("/{producto_id}", response_model=ProductoExtraOut)
def actualizar_producto(
    producto_id: int, data: ProductoExtraUpdate,
    db: Session = Depends(get_db), _=Depends(require_admin),
):
    producto = db.get(ProductoConExtra, producto_id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(producto, key, value)
    db.commit()
    _recalcular_extras_envios(db, {producto.codigo_mlc})
    db.commit()
    db.refresh(producto)
    return producto


@router.delete("/{producto_id}")
def eliminar_producto(producto_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    producto = db.get(ProductoConExtra, producto_id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    producto.activo = False
    db.commit()
    return {"message": "Producto desactivado"}
