"""
Endpoints del portal de seller y driver:
- Liquidación propia (seller y driver)
- Facturación propia (seller)
- PDF filtrado (seller sin datos driver, driver sin datos seller)
- Excel filtrado (mismas restricciones)
"""
import io
import os
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth import require_seller, require_driver, driver_period_allowed
from app.models import (
    Envio, Seller, Driver, Retiro, AjusteLiquidacion,
    TipoEntidadEnum, PagoSemanaSeller, FacturaMensualSeller,
    CalendarioSemanas, EstadoPagoEnum, PagoCartolaSeller,
    PagoSemanaDriver, FacturaDriver, EstadoFacturaDriverEnum,
)
from app.services.liquidacion import calcular_liquidacion_sellers, calcular_liquidacion_drivers
from app.services.pdf_generator import generar_pdf_seller, generar_pdf_driver
from app.api.liquidacion import _driver_detail, _daily_breakdown, _productos_envios

UPLOADS_DIR_DRIVERS = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "uploads", "facturas_drivers",
)


def _fmt_fecha(valor) -> str:
    """Normaliza cualquier valor de fecha a string ISO yyyy-mm-dd."""
    if not valor:
        return ""
    from datetime import date as _date, datetime as _dt
    if isinstance(valor, (_date, _dt)):
        return valor.strftime("%Y-%m-%d")
    s = str(valor).strip()
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 3:
            if len(parts[0]) <= 2:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
            else:
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    return s

router = APIRouter(prefix="/portal", tags=["Portal"])

MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _semanas_del_mes(db: Session, mes: int, anio: int):
    rows = db.query(CalendarioSemanas.semana).filter(
        CalendarioSemanas.mes == mes,
        CalendarioSemanas.anio == anio,
    ).order_by(CalendarioSemanas.semana).all()
    return [r[0] for r in rows] if rows else [1, 2, 3, 4, 5]


def _monto_semanal_seller(db, seller_id, semana, mes, anio):
    envios = db.query(Envio).filter(
        Envio.seller_id == seller_id, Envio.semana == semana,
        Envio.mes == mes, Envio.anio == anio,
    ).all()
    if not envios:
        return 0
    seller = db.get(Seller, seller_id)
    total = sum(e.cobro_seller + e.cobro_extra_manual + e.extra_producto_seller + e.extra_comuna_seller for e in envios)
    if seller and seller.tiene_retiro and not seller.usa_pickup:
        if not (seller.min_paquetes_retiro_gratis > 0 and len(envios) >= seller.min_paquetes_retiro_gratis):
            retiros = db.query(Retiro).filter(
                Retiro.seller_id == seller_id, Retiro.semana == semana,
                Retiro.mes == mes, Retiro.anio == anio,
            ).all()
            if retiros:
                total += sum(r.tarifa_seller for r in retiros)
            elif seller.tarifa_retiro and (semana >= 4 if (mes == 2 and anio == 2026) else (anio, mes) > (2026, 2)):
                dias_con_envios = len({e.fecha_entrega for e in envios if e.fecha_entrega})
                total += seller.tarifa_retiro * dias_con_envios
    ajustes = db.query(AjusteLiquidacion).filter(
        AjusteLiquidacion.tipo == TipoEntidadEnum.SELLER,
        AjusteLiquidacion.entidad_id == seller_id,
        AjusteLiquidacion.semana == semana,
        AjusteLiquidacion.mes == mes,
        AjusteLiquidacion.anio == anio,
    ).all()
    total += sum(a.monto for a in ajustes)
    return total


# ── SELLER ──────────────────────────────────────────────────────────────────

@router.get("/seller/liquidacion")
def seller_liquidacion(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_seller),
):
    seller_id = current_user["id"]
    resultado = calcular_liquidacion_sellers(db, semana, mes, anio)
    row = next((r for r in resultado if r["seller_id"] == seller_id), None)
    if not row:
        return {
            "seller_id": seller_id, "semana": semana, "mes": mes, "anio": anio,
            "cantidad_envios": 0, "total_envios": 0,
            "total_extras_producto": 0, "total_extras_comuna": 0,
            "total_retiros": 0, "total_ajustes": 0,
            "subtotal": 0, "iva": 0, "total_con_iva": 0,
        }
    # Excluir campos de driver del resultado
    return {k: v for k, v in row.items() if k not in ("empresa", "user_nombres")}


@router.get("/seller/facturacion")
def seller_facturacion(
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_seller),
):
    seller_id = current_user["id"]
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(status_code=404)

    semanas = _semanas_del_mes(db, mes, anio)
    semanas_data = {}
    subtotal = 0

    for sem in semanas:
        pago = db.query(PagoSemanaSeller).filter(
            PagoSemanaSeller.seller_id == seller_id,
            PagoSemanaSeller.semana == sem,
            PagoSemanaSeller.mes == mes,
            PagoSemanaSeller.anio == anio,
        ).first()
        monto = pago.monto_override if (pago and pago.monto_override is not None) else _monto_semanal_seller(db, seller_id, sem, mes, anio)
        estado = pago.estado if pago else EstadoPagoEnum.PENDIENTE.value
        semanas_data[str(sem)] = {"monto_neto": monto, "estado": estado}
        subtotal += monto

    iva = int(subtotal * 0.19)
    factura = db.query(FacturaMensualSeller).filter(
        FacturaMensualSeller.seller_id == seller_id,
        FacturaMensualSeller.mes == mes,
        FacturaMensualSeller.anio == anio,
    ).first()

    return {
        "seller_id": seller_id,
        "seller_nombre": seller.nombre,
        "rut": seller.rut,
        "mes": mes, "anio": anio,
        "semanas_disponibles": semanas,
        "semanas": semanas_data,
        "subtotal_neto": subtotal,
        "iva": iva,
        "total_con_iva": subtotal + iva,
        "factura_estado": factura.estado if factura else None,
        "factura_folio": factura.folio_haulmer if factura else None,
    }


@router.get("/seller/facturas-historial")
def seller_facturas_historial(
    limite: int = Query(24, ge=1, le=60),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_seller),
):
    """Historial de facturas emitidas para el seller actual."""
    from app.models import EstadoFacturaEnum
    seller_id = current_user["id"]
    facturas = (
        db.query(FacturaMensualSeller)
        .filter(
            FacturaMensualSeller.seller_id == seller_id,
            FacturaMensualSeller.estado == EstadoFacturaEnum.EMITIDA.value,
        )
        .order_by(FacturaMensualSeller.anio.desc(), FacturaMensualSeller.mes.desc())
        .limit(limite)
        .all()
    )
    return [
        {
            "mes": f.mes,
            "anio": f.anio,
            "total": f.total,
            "folio_haulmer": f.folio_haulmer,
            "emitida_en": f.emitida_en.isoformat() if f.emitida_en else None,
        }
        for f in facturas
    ]


@router.get("/seller/pdf")
def seller_pdf(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_seller),
):
    seller_id = current_user["id"]
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(status_code=404)
    try:
        pdf_bytes = generar_pdf_seller(db, seller_id, semana, mes, anio)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Sin datos para este período: {e}")
    nombre = seller.nombre.replace("/", "-")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=liquidacion_{nombre}_S{semana}_M{mes}_{anio}.pdf"},
    )


@router.get("/seller/excel")
def seller_excel(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_seller),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    seller_id = current_user["id"]
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(status_code=404)

    envios = db.query(Envio).filter(
        Envio.seller_id == seller_id,
        Envio.semana == semana, Envio.mes == mes, Envio.anio == anio,
    ).order_by(Envio.fecha_entrega).all()

    if not envios:
        raise HTTPException(status_code=404, detail="Sin envíos para este período")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mis Envíos"

    # Solo columnas visibles para seller (sin info de driver)
    headers = [
        "Fecha Entrega", "Tracking", "Seller Code", "External ID",
        "Comuna", "Bultos",
        "Descripción Producto", "Código MLC",
        "Cobro Base", "Extra Prod.", "Extra Com.", "Extra Manual", "Total",
    ]
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        c.border = thin

    for idx, e in enumerate(envios, 2):
        total = e.cobro_seller + e.extra_producto_seller + e.extra_comuna_seller + (e.cobro_extra_manual or 0)
        row = [
            str(e.fecha_entrega) if e.fecha_entrega else "",
            e.tracking_id or "",
            e.seller_code or "",
            e.venta_id or "",
            (e.comuna or "").title(),
            e.bultos,
            e.descripcion_producto or "",
            e.codigo_producto or "",
            e.cobro_seller,
            e.extra_producto_seller,
            e.extra_comuna_seller,
            e.cobro_extra_manual or 0,
            total,
        ]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=idx, column=ci, value=v)
            c.border = thin

    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
    ws.freeze_panes = "A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")), 10) + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = seller.nombre.replace("/", "-").replace("\\", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=envios_{nombre}_S{semana}_M{mes}_{anio}.xlsx"},
    )


# ── SELLER: Ganancias ────────────────────────────────────────────────────────

@router.get("/seller/ganancias")
def seller_ganancias(
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_seller),
):
    """
    Resumen de pagos recibidos del seller autenticado para un mes/año.
    - Tabla semanal: liquidado, cobrado, estado por semana.
    - Historial de pagos individuales (PagoCartolaSeller) del mes.
    """
    from app.api.facturacion import _get_monto_semanal_seller

    seller_id = current_user["id"]
    seller = db.get(Seller, seller_id)
    if not seller:
        raise HTTPException(status_code=404, detail="Seller no encontrado")

    semanas_rows = db.query(CalendarioSemanas.semana).filter(
        CalendarioSemanas.mes == mes,
        CalendarioSemanas.anio == anio,
    ).order_by(CalendarioSemanas.semana).all()
    semanas = [r[0] for r in semanas_rows] if semanas_rows else [1, 2, 3, 4, 5]

    # ── Tabla semanal ────────────────────────────────────────────────────────
    semanas_detalle = []
    total_liquidado = 0
    total_cobrado = 0

    for sem in semanas:
        liquidado = _get_monto_semanal_seller(db, seller_id, sem, mes, anio)
        if liquidado == 0:
            continue

        pago = db.query(PagoSemanaSeller).filter(
            PagoSemanaSeller.seller_id == seller_id,
            PagoSemanaSeller.semana == sem,
            PagoSemanaSeller.mes == mes,
            PagoSemanaSeller.anio == anio,
        ).first()

        estado = pago.estado if pago else EstadoPagoEnum.PENDIENTE.value
        cobrado = liquidado if estado == EstadoPagoEnum.PAGADO.value else (
            pago.monto_neto if pago and estado == EstadoPagoEnum.INCOMPLETO.value else 0
        )

        semanas_detalle.append({
            "semana": sem,
            "liquidado": liquidado,
            "cobrado": cobrado,
            "estado": estado,
        })
        total_liquidado += liquidado
        total_cobrado += cobrado

    # ── Pagos individuales (PagoCartolaSeller) ───────────────────────────────
    pagos_rows = db.query(PagoCartolaSeller).filter(
        PagoCartolaSeller.seller_id == seller_id,
        PagoCartolaSeller.mes == mes,
        PagoCartolaSeller.anio == anio,
    ).order_by(PagoCartolaSeller.fecha_pago.desc(), PagoCartolaSeller.created_at.desc()).all()

    pagos = [
        {
            "id": p.id,
            "fecha_pago": _fmt_fecha(p.fecha_pago),
            "semana": p.semana,
            "monto": p.monto,
            "fuente": p.fuente,
            "descripcion": p.descripcion,
        }
        for p in pagos_rows
    ]

    return {
        "mes": mes,
        "anio": anio,
        "resumen": {
            "total_liquidado": total_liquidado,
            "total_cobrado": total_cobrado,
            "pendiente": total_liquidado - total_cobrado,
        },
        "semanas": semanas_detalle,
        "pagos": pagos,
    }


# ── DRIVER ──────────────────────────────────────────────────────────────────

@router.get("/driver/liquidacion")
def driver_liquidacion(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    sub_driver_id: Optional[int] = Query(None, description="ID de subordinado (solo jefes de flota)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    if not driver_period_allowed(anio, mes, semana):
        raise HTTPException(
            status_code=403,
            detail="Solo puedes ver información desde la semana 4 de febrero 2026 en adelante.",
        )
    driver_id = current_user["id"]

    # Si se pide un subordinado, verificar que el solicitante es su jefe
    if sub_driver_id and sub_driver_id != driver_id:
        sub = db.query(Driver).filter(
            Driver.id == sub_driver_id,
            Driver.jefe_flota_id == driver_id,
            Driver.activo == True,
        ).first()
        if not sub:
            raise HTTPException(status_code=403, detail="No tienes acceso a la liquidación de este conductor.")
        target_id = sub_driver_id
    else:
        target_id = driver_id

    detail = _driver_detail(db, target_id, mes, anio)
    driver = db.get(Driver, target_id)
    es_contratado = getattr(driver, 'contratado', False) if driver else False
    envios_semana = db.query(Envio).filter(
        Envio.driver_id == target_id, Envio.semana == semana,
        Envio.mes == mes, Envio.anio == anio,
    ).order_by(Envio.fecha_entrega).all()
    retiros_semana = db.query(Retiro).filter(
        Retiro.driver_id == target_id, Retiro.semana == semana,
        Retiro.mes == mes, Retiro.anio == anio,
    ).all()
    detail["daily"] = _daily_breakdown(
        envios_semana, retiros_semana,
        "extra_producto_driver", "extra_comuna_driver",
        semana, mes, anio, is_seller=False, db=db, contratado=es_contratado, driver=driver,
    )
    detail["productos"] = [] if es_contratado else _productos_envios(db, envios_semana)

    # Incluir lista de subordinados si el driver autenticado es jefe
    subordinados = db.query(Driver).filter(
        Driver.jefe_flota_id == driver_id,
        Driver.activo == True,
    ).order_by(Driver.nombre).all()
    detail["es_jefe"] = len(subordinados) > 0
    detail["subordinados"] = [{"id": s.id, "nombre": s.nombre} for s in subordinados]
    detail["conductor_id"] = target_id
    detail["conductor_nombre"] = driver.nombre if driver else ""

    return detail


@router.get("/driver/excel")
def driver_excel(
    semana: Optional[int] = None,
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    if semana is not None and not driver_period_allowed(anio, mes, semana):
        raise HTTPException(
            status_code=403,
            detail="Solo puedes descargar información desde la semana 4 de febrero 2026 en adelante.",
        )
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    driver_id = current_user["id"]
    driver = db.get(Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404)

    query = db.query(Envio).filter(
        Envio.driver_id == driver_id,
        Envio.mes == mes, Envio.anio == anio,
    )
    if semana is not None:
        query = query.filter(Envio.semana == semana)
    envios = query.order_by(Envio.fecha_entrega).all()

    if not envios:
        raise HTTPException(status_code=404, detail="Sin entregas para este período")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mis Entregas"

    es_contratado = getattr(driver, 'contratado', False)
    if es_contratado:
        headers = [
            "Fecha Entrega", "Tracking", "Seller Code", "External ID", "Seller", "Comuna", "Bultos",
            "Descripción Producto",
            "Pago Base", "Pago Extra Manual", "Total",
        ]
    else:
        headers = [
            "Fecha Entrega", "Tracking", "Seller Code", "External ID", "Seller", "Comuna", "Bultos",
            "Descripción Producto",
            "Pago Base", "Extra Prod.", "Extra Com.", "Pago Extra Manual", "Total",
        ]
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        c.border = thin

    for idx, e in enumerate(envios, 2):
        seller = db.get(Seller, e.seller_id) if e.seller_id else None
        if es_contratado:
            total = e.costo_driver + (e.pago_extra_manual or 0)
            row = [
                str(e.fecha_entrega) if e.fecha_entrega else "",
                e.tracking_id or "",
                e.seller_code or "",
                e.venta_id or "",
                seller.nombre if seller else (e.seller_nombre_raw or ""),
                (e.comuna or "").title(),
                e.bultos,
                e.descripcion_producto or "",
                e.costo_driver,
                e.pago_extra_manual or 0,
                total,
            ]
        else:
            total = e.costo_driver + e.extra_producto_driver + e.extra_comuna_driver + (e.pago_extra_manual or 0)
            row = [
                str(e.fecha_entrega) if e.fecha_entrega else "",
                e.tracking_id or "",
                e.seller_code or "",
                e.venta_id or "",
                seller.nombre if seller else (e.seller_nombre_raw or ""),
                (e.comuna or "").title(),
                e.bultos,
                e.descripcion_producto or "",
                e.costo_driver,
                e.extra_producto_driver,
                e.extra_comuna_driver,
                e.pago_extra_manual or 0,
                total,
            ]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=idx, column=ci, value=v)
            c.border = thin

    last_envio_row = len(envios) + 1

    ajuste_query = db.query(AjusteLiquidacion).filter(
        AjusteLiquidacion.tipo == TipoEntidadEnum.DRIVER,
        AjusteLiquidacion.entidad_id == driver_id,
        AjusteLiquidacion.mes == mes,
        AjusteLiquidacion.anio == anio,
    )
    if semana is not None:
        ajuste_query = ajuste_query.filter(AjusteLiquidacion.semana == semana)
    ajustes = ajuste_query.order_by(AjusteLiquidacion.semana).all()

    if ajustes:
        ajuste_start = last_envio_row + 2
        ajuste_hfill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        ajuste_headers = ["Semana", "Tipo", "Motivo", "Monto"]
        for i, h in enumerate(ajuste_headers, 1):
            c = ws.cell(row=ajuste_start, column=i, value=h)
            c.font = hfont
            c.fill = ajuste_hfill
            c.border = thin

        for ai, a in enumerate(ajustes, ajuste_start + 1):
            tipo_label = "Bonificación" if a.monto > 0 else "Descuento"
            for ci, v in enumerate([a.semana, tipo_label, a.motivo or "", a.monto], 1):
                c = ws.cell(row=ai, column=ci, value=v)
                c.border = thin

    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
    ws.freeze_panes = "A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")), 10) + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = driver.nombre.replace("/", "-").replace("\\", "-")
    period_suffix = f"_S{semana}_M{mes}_{anio}" if semana is not None else f"_M{mes}_{anio}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=entregas_{nombre}{period_suffix}.xlsx"},
    )


@router.get("/driver/pagos-recibidos")
def driver_pagos_recibidos(
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    """
    Devuelve los pagos recibidos (semanas en estado PAGADO) para el driver autenticado.
    Incluye monto liquidado y monto pagado por semana.
    """
    from app.models import PagoSemanaDriver
    from app.api.cpc import _get_monto_semanal_driver

    driver_id = current_user["id"]

    semanas_rows = db.query(CalendarioSemanas.semana).filter(
        CalendarioSemanas.mes == mes,
        CalendarioSemanas.anio == anio,
    ).order_by(CalendarioSemanas.semana).all()
    semanas = [r[0] for r in semanas_rows] if semanas_rows else [1, 2, 3, 4, 5]

    resultado = []
    total_liquidado = 0
    total_pagado = 0

    for sem in semanas:
        liquidado = _get_monto_semanal_driver(db, driver_id, sem, mes, anio)
        if liquidado == 0:
            continue

        pago = db.query(PagoSemanaDriver).filter(
            PagoSemanaDriver.driver_id == driver_id,
            PagoSemanaDriver.semana == sem,
            PagoSemanaDriver.mes == mes,
            PagoSemanaDriver.anio == anio,
        ).first()

        estado = pago.estado if pago else EstadoPagoEnum.PENDIENTE.value
        pagado = liquidado if estado == EstadoPagoEnum.PAGADO.value else 0

        resultado.append({
            "semana": sem,
            "liquidado": liquidado,
            "pagado": pagado,
            "estado": estado,
        })
        total_liquidado += liquidado
        total_pagado += pagado

    return {
        "mes": mes,
        "anio": anio,
        "semanas": resultado,
        "total_liquidado": total_liquidado,
        "total_pagado": total_pagado,
    }


@router.get("/driver/ganancias")
def driver_ganancias(
    mes: int = Query(...),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    """
    Resumen de ganancias del driver autenticado para un mes/año.
    - Tabla semanal: liquidado, pagado, estado por semana.
    - Historial de pagos individuales (PagoCartola) del mes.
    Si es jefe de flota, incluye también las semanas y pagos de sus subordinados
    (señalizados con el nombre del driver).
    """
    from app.models import PagoSemanaDriver, PagoCartola
    from app.api.cpc import _get_monto_semanal_driver

    driver_id = current_user["id"]
    driver = db.get(Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")

    subordinados = db.query(Driver).filter(
        Driver.jefe_flota_id == driver_id,
        Driver.activo == True,
    ).all()
    es_jefe = len(subordinados) > 0

    semanas_rows = db.query(CalendarioSemanas.semana).filter(
        CalendarioSemanas.mes == mes,
        CalendarioSemanas.anio == anio,
    ).order_by(CalendarioSemanas.semana).all()
    semanas = [r[0] for r in semanas_rows] if semanas_rows else [1, 2, 3, 4, 5]

    # ── Tabla semanal ────────────────────────────────────────────────────────
    drivers_a_calcular = [{"id": driver_id, "nombre": driver.nombre, "es_propio": True}]
    if es_jefe:
        for sub in subordinados:
            drivers_a_calcular.append({"id": sub.id, "nombre": sub.nombre, "es_propio": False})

    semanas_detalle = []
    total_liquidado = 0
    total_pagado = 0

    for sem in semanas:
        for d in drivers_a_calcular:
            liquidado = _get_monto_semanal_driver(db, d["id"], sem, mes, anio)
            if liquidado == 0:
                continue

            pago = db.query(PagoSemanaDriver).filter(
                PagoSemanaDriver.driver_id == d["id"],
                PagoSemanaDriver.semana == sem,
                PagoSemanaDriver.mes == mes,
                PagoSemanaDriver.anio == anio,
            ).first()

            estado = pago.estado if pago else EstadoPagoEnum.PENDIENTE.value
            pagado = liquidado if estado == EstadoPagoEnum.PAGADO.value else (
                pago.monto_neto if pago and estado == EstadoPagoEnum.INCOMPLETO.value else 0
            )

            semanas_detalle.append({
                "semana": sem,
                "driver_id": d["id"],
                "driver_nombre": d["nombre"],
                "es_propio": d["es_propio"],
                "liquidado": liquidado,
                "pagado": pagado,
                "estado": estado,
            })

            if d["es_propio"] or not es_jefe:
                total_liquidado += liquidado
                total_pagado += pagado

    # ── Pagos individuales (PagoCartola) ────────────────────────────────────
    ids_a_buscar = [driver_id] + ([s.id for s in subordinados] if es_jefe else [])

    pagos_rows = db.query(PagoCartola).filter(
        PagoCartola.driver_id.in_(ids_a_buscar),
        PagoCartola.mes == mes,
        PagoCartola.anio == anio,
    ).order_by(PagoCartola.fecha_pago.desc(), PagoCartola.created_at.desc()).all()

    # Mapa id→nombre para subordinados
    nombre_map = {d["id"]: d["nombre"] for d in drivers_a_calcular}

    pagos = [
        {
            "id": p.id,
            "fecha_pago": _fmt_fecha(p.fecha_pago),
            "semana": p.semana,
            "monto": p.monto,
            "fuente": p.fuente,
            "descripcion": p.descripcion,
            "driver_id": p.driver_id,
            "driver_nombre": nombre_map.get(p.driver_id, ""),
            "es_propio": p.driver_id == driver_id,
        }
        for p in pagos_rows
    ]

    return {
        "mes": mes,
        "anio": anio,
        "es_jefe": es_jefe,
        "resumen": {
            "total_liquidado": total_liquidado,
            "total_pagado": total_pagado,
            "pendiente": total_liquidado - total_pagado,
        },
        "semanas": semanas_detalle,
        "pagos": pagos,
    }


# ── DRIVER: Facturas semanales ────────────────────────────────────────────────

def _monto_efectivo_driver(db: Session, driver_id: int, semana: int, mes: int, anio: int) -> int:
    """Retorna el monto exactamente como lo muestra CPC para esa semana."""
    from app.api.cpc import _get_monto_semanal_driver
    pago = db.query(PagoSemanaDriver).filter(
        PagoSemanaDriver.driver_id == driver_id,
        PagoSemanaDriver.semana == semana,
        PagoSemanaDriver.mes == mes,
        PagoSemanaDriver.anio == anio,
    ).first()
    if pago and pago.monto_override is not None:
        return pago.monto_override
    if pago and pago.estado == EstadoPagoEnum.PAGADO.value and pago.monto_neto:
        return pago.monto_neto
    return _get_monto_semanal_driver(db, driver_id, semana, mes, anio)


@router.post("/driver/facturas/upload")
async def driver_upload_factura(
    semana: int = Query(...),
    mes: int = Query(...),
    anio: int = Query(...),
    nota: Optional[str] = Query(None),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    driver_id = current_user["id"]
    driver = db.get(Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")

    allowed_ext = (".pdf", ".jpg", ".jpeg", ".png", ".webp")
    ext = os.path.splitext(archivo.filename or "")[1].lower()
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail=f"Formato no permitido. Use: {', '.join(allowed_ext)}")

    existing = db.query(FacturaDriver).filter(
        FacturaDriver.driver_id == driver_id,
        FacturaDriver.semana == semana,
        FacturaDriver.mes == mes,
        FacturaDriver.anio == anio,
    ).first()

    if existing and existing.estado == EstadoFacturaDriverEnum.APROBADA.value:
        raise HTTPException(status_code=400, detail="La factura de esta semana ya fue aprobada")

    os.makedirs(UPLOADS_DIR_DRIVERS, exist_ok=True)
    unique_name = f"{driver_id}_{semana}_{mes}_{anio}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = os.path.join(UPLOADS_DIR_DRIVERS, unique_name)

    content = await archivo.read()
    with open(file_path, "wb") as f:
        f.write(content)

    if existing:
        if existing.archivo_path and os.path.exists(existing.archivo_path):
            try:
                os.remove(existing.archivo_path)
            except OSError:
                pass
        existing.archivo_nombre = archivo.filename
        existing.archivo_path = file_path
        existing.estado = EstadoFacturaDriverEnum.CARGADA.value
        existing.nota_driver = nota
        existing.nota_admin = None
        existing.revisado_por = None
        existing.revisado_en = None
    else:
        existing = FacturaDriver(
            driver_id=driver_id, semana=semana, mes=mes, anio=anio,
            archivo_nombre=archivo.filename,
            archivo_path=file_path,
            estado=EstadoFacturaDriverEnum.CARGADA.value,
            nota_driver=nota,
        )
        db.add(existing)

    db.commit()
    db.refresh(existing)

    return {
        "ok": True,
        "factura_id": existing.id,
        "estado": existing.estado,
        "archivo_nombre": existing.archivo_nombre,
    }


@router.get("/driver/facturas")
def driver_listar_facturas(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    driver_id = current_user["id"]
    facturas = (
        db.query(FacturaDriver)
        .filter(FacturaDriver.driver_id == driver_id)
        .order_by(FacturaDriver.anio.desc(), FacturaDriver.mes.desc(), FacturaDriver.semana.desc())
        .all()
    )
    return [
        {
            "id": f.id,
            "semana": f.semana,
            "mes": f.mes,
            "anio": f.anio,
            "monto_neto": _monto_efectivo_driver(db, driver_id, f.semana, f.mes, f.anio),
            "archivo_nombre": f.archivo_nombre,
            "estado": f.estado,
            "nota_driver": f.nota_driver,
            "nota_admin": f.nota_admin,
            "revisado_por": f.revisado_por,
            "revisado_en": f.revisado_en.isoformat() if f.revisado_en else None,
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in facturas
    ]


@router.get("/driver/facturas/{factura_id}/descargar")
def driver_descargar_factura(
    factura_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_driver),
):
    factura = db.get(FacturaDriver, factura_id)
    if not factura or factura.driver_id != current_user["id"]:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if not factura.archivo_path or not os.path.exists(factura.archivo_path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(
        factura.archivo_path,
        filename=factura.archivo_nombre or "factura.pdf",
        media_type="application/octet-stream",
    )
