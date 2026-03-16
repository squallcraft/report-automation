from typing import Optional, List
from difflib import SequenceMatcher

import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pydantic import BaseModel
import pandas as pd

from app.database import get_db
from app.auth import require_permission
from app.models import Retiro, Seller, Driver, Pickup, Sucursal, PagoSemanaDriver, PagoSemanaSeller, PagoSemanaPickup
from app.schemas import RetiroCreate, RetiroUpdate, RetiroOut
from app.services.ingesta import calcular_semana_del_mes, homologar_nombre
from app.services.audit import registrar as audit
from app.api.liquidacion import invalidar_snapshots
from app.models import EstadoPagoEnum


def _similaridad(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _mejor_match(nombre_raw: str, entidades) -> tuple:
    """Busca la mejor coincidencia entre nombre_raw y una lista de entidades (con .nombre y .aliases)."""
    mejor, mejor_score = None, 0.0
    nombre_norm = nombre_raw.lower().strip()
    for ent in entidades:
        score = _similaridad(nombre_norm, ent.nombre.lower())
        for alias in (ent.aliases or []):
            s = _similaridad(nombre_norm, alias.lower())
            if s > score:
                score = s
        if score > mejor_score:
            mejor_score = score
            mejor = ent
    return mejor, round(mejor_score, 2)

router = APIRouter(prefix="/retiros", tags=["Retiros"])


def _verificar_semana_abierta(db: Session, semana: int, mes: int, anio: int,
                               driver_id: int = None, seller_id: int = None, pickup_id: int = None):
    """Lanza HTTP 409 si la semana ya está cerrada (pago en estado PAGADO) para el driver/seller/pickup del retiro."""
    if driver_id:
        psd = db.query(PagoSemanaDriver).filter(
            PagoSemanaDriver.driver_id == driver_id,
            PagoSemanaDriver.semana == semana,
            PagoSemanaDriver.mes == mes,
            PagoSemanaDriver.anio == anio,
            PagoSemanaDriver.estado == EstadoPagoEnum.PAGADO.value,
        ).first()
        if psd:
            raise HTTPException(
                status_code=409,
                detail=f"La semana {semana} ({mes}/{anio}) ya está cerrada para este conductor. No se pueden modificar retiros en semanas pagadas.",
            )
    if seller_id:
        pss = db.query(PagoSemanaSeller).filter(
            PagoSemanaSeller.seller_id == seller_id,
            PagoSemanaSeller.semana == semana,
            PagoSemanaSeller.mes == mes,
            PagoSemanaSeller.anio == anio,
            PagoSemanaSeller.estado == EstadoPagoEnum.PAGADO.value,
        ).first()
        if pss:
            raise HTTPException(
                status_code=409,
                detail=f"La semana {semana} ({mes}/{anio}) ya está cerrada para este seller. No se pueden modificar retiros en semanas pagadas.",
            )
    if pickup_id:
        psp = db.query(PagoSemanaPickup).filter(
            PagoSemanaPickup.pickup_id == pickup_id,
            PagoSemanaPickup.semana == semana,
            PagoSemanaPickup.mes == mes,
            PagoSemanaPickup.anio == anio,
            PagoSemanaPickup.estado == EstadoPagoEnum.PAGADO.value,
        ).first()
        if psp:
            raise HTTPException(
                status_code=409,
                detail=f"La semana {semana} ({mes}/{anio}) ya está cerrada para este pickup. No se pueden modificar retiros en semanas pagadas.",
            )


def _check_cerrado(db: Session, semana: int, mes: int, anio: int,
                   driver_id: int = None, seller_id: int = None, pickup_id: int = None) -> bool:
    """Versión booleana de _verificar_semana_abierta, para uso en validación batch."""
    if driver_id:
        if db.query(PagoSemanaDriver).filter(
            PagoSemanaDriver.driver_id == driver_id,
            PagoSemanaDriver.semana == semana,
            PagoSemanaDriver.mes == mes,
            PagoSemanaDriver.anio == anio,
            PagoSemanaDriver.estado == EstadoPagoEnum.PAGADO.value,
        ).first():
            return True
    if seller_id:
        if db.query(PagoSemanaSeller).filter(
            PagoSemanaSeller.seller_id == seller_id,
            PagoSemanaSeller.semana == semana,
            PagoSemanaSeller.mes == mes,
            PagoSemanaSeller.anio == anio,
            PagoSemanaSeller.estado == EstadoPagoEnum.PAGADO.value,
        ).first():
            return True
    if pickup_id:
        if db.query(PagoSemanaPickup).filter(
            PagoSemanaPickup.pickup_id == pickup_id,
            PagoSemanaPickup.semana == semana,
            PagoSemanaPickup.mes == mes,
            PagoSemanaPickup.anio == anio,
            PagoSemanaPickup.estado == EstadoPagoEnum.PAGADO.value,
        ).first():
            return True
    return False


@router.get("/semanas-cerradas")
def semanas_cerradas(
    mes: int,
    anio: int,
    db: Session = Depends(get_db),
    _=require_permission("retiros:ver"),
):
    """
    Retorna qué (driver_id, semana), (seller_id, semana) y (pickup_id, semana) están cerradas (PAGADO)
    para el mes/año dado. El frontend usa esto para bloquear acciones sobre retiros cerrados.
    """
    drivers_cerrados: dict[str, list[int]] = {}
    sellers_cerrados: dict[str, list[int]] = {}
    pickups_cerrados: dict[str, list[int]] = {}

    for psd in db.query(PagoSemanaDriver).filter(
        PagoSemanaDriver.mes == mes,
        PagoSemanaDriver.anio == anio,
        PagoSemanaDriver.estado == EstadoPagoEnum.PAGADO.value,
    ).all():
        drivers_cerrados.setdefault(str(psd.driver_id), []).append(psd.semana)

    for pss in db.query(PagoSemanaSeller).filter(
        PagoSemanaSeller.mes == mes,
        PagoSemanaSeller.anio == anio,
        PagoSemanaSeller.estado == EstadoPagoEnum.PAGADO.value,
    ).all():
        sellers_cerrados.setdefault(str(pss.seller_id), []).append(pss.semana)

    for psp in db.query(PagoSemanaPickup).filter(
        PagoSemanaPickup.mes == mes,
        PagoSemanaPickup.anio == anio,
        PagoSemanaPickup.estado == EstadoPagoEnum.PAGADO.value,
    ).all():
        pickups_cerrados.setdefault(str(psp.pickup_id), []).append(psp.semana)

    return {
        "drivers": drivers_cerrados,
        "sellers": sellers_cerrados,
        "pickups": pickups_cerrados,
    }


def _enrich_retiro(r, db) -> dict:
    data = {col.name: getattr(r, col.name) for col in r.__table__.columns}
    seller = db.get(Seller, r.seller_id) if r.seller_id else None
    driver = db.get(Driver, r.driver_id) if r.driver_id else None
    pickup = db.get(Pickup, r.pickup_id) if r.pickup_id else None
    sucursal = db.get(Sucursal, r.sucursal_id) if r.sucursal_id else None
    data["seller_nombre"] = seller.nombre if seller else r.seller_nombre_raw or "—"
    data["driver_nombre"] = driver.nombre if driver else r.driver_nombre_raw or "—"
    data["pickup_nombre"] = pickup.nombre if pickup else None
    data["sucursal_nombre"] = sucursal.nombre if sucursal else None
    return data


@router.get("", response_model=List[RetiroOut])
def listar_retiros(
    semana: Optional[int] = None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
    db: Session = Depends(get_db),
    _=require_permission("retiros:ver"),
):
    query = db.query(Retiro)
    if semana is not None:
        query = query.filter(Retiro.semana == semana)
    if mes is not None:
        query = query.filter(Retiro.mes == mes)
    if anio is not None:
        query = query.filter(Retiro.anio == anio)
    retiros = query.order_by(Retiro.fecha.desc()).all()
    return [_enrich_retiro(r, db) for r in retiros]


@router.get("/plantilla/descargar")
def descargar_plantilla_retiros():
    """Genera plantilla Excel para importación masiva de retiros."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Retiros"

    columnas = ["Fecha Retiro", "Nombre Conductor", "Nombre Seller"]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    for i, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = border

    ejemplos = [
        ["2026-02-28", "Carlos", "MercadoLibre Chile"],
        ["2026-02-28", "Augusto", "Ferretería Oviedo"],
        ["2026-02-27", "Jorge", "Aventura Store"],
    ]

    alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    for ri, fila in enumerate(ejemplos, 2):
        for ci, val in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=10)
            c.border = border
            if ri % 2 == 0:
                c.fill = alt

    anchos = [18, 28, 30]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.auto_filter.ref = "A1:C1"
    ws.freeze_panes = "A2"

    wsi = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("Instrucciones para Importación de Retiros", ""),
        ("", ""),
        ("Campo", "Descripción"),
        ("Fecha Retiro", "Fecha en formato YYYY-MM-DD o DD/MM/YYYY. El sistema calcula automáticamente semana, mes y año."),
        ("Nombre Conductor", "Nombre del conductor que realizó el retiro. Se homologa automáticamente con los drivers registrados."),
        ("Nombre Seller", "Nombre del seller donde se hizo el retiro. Se homologa con los sellers registrados. Las tarifas se obtienen automáticamente del seller."),
        ("", ""),
        ("Notas importantes:", ""),
        ("", "Los sellers con 'usa_pickup = Sí' serán ignorados (pickup no genera retiro)."),
        ("", "La tarifa de cobro al seller y el pago al driver se toman de la configuración del seller."),
        ("", "Si un nombre no se puede homologar, el retiro queda como 'sin homologar' y se puede resolver manualmente."),
    ]
    for ri, (a, b) in enumerate(instrucciones, 1):
        ca = wsi.cell(row=ri, column=1, value=a)
        cb = wsi.cell(row=ri, column=2, value=b)
        if ri == 1:
            ca.font = Font(bold=True, size=14, color="1A365D")
        elif ri == 3:
            ca.font = Font(bold=True, size=11)
            cb.font = Font(bold=True, size=11)
        else:
            ca.font = Font(bold=True, size=10) if a else Font(size=10)
            cb.font = Font(size=10, color="4A5568")
    wsi.column_dimensions["A"].width = 22
    wsi.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_retiros.xlsx"})


@router.post("/importar/preview")
async def preview_retiros(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=require_permission("retiros:editar"),
):
    """Parsea Excel de retiros y retorna un preview con matches propuestos. No graba nada."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx / .xls)")

    content = await file.read()
    df = pd.read_excel(io.BytesIO(content))

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "fecha" in cl:
            col_map[col] = "fecha"
        elif "conductor" in cl or "driver" in cl:
            col_map[col] = "conductor"
        elif "seller" in cl or "cliente" in cl:
            col_map[col] = "seller"
    df = df.rename(columns=col_map)

    sellers = db.query(Seller).filter(Seller.activo == True).all()
    drivers = db.query(Driver).filter(Driver.activo == True).all()
    pickups_list = db.query(Pickup).filter(Pickup.activo == True).all()
    sucursales_list = db.query(Sucursal).filter(Sucursal.activo == True).all()

    items = []
    errores = []

    for idx, row in df.iterrows():
        try:
            fecha_raw = row.get("fecha")
            if pd.isna(fecha_raw):
                errores.append(f"Fila {idx + 2}: sin fecha")
                continue
            if isinstance(fecha_raw, str):
                fecha = pd.to_datetime(fecha_raw).date()
            else:
                fecha = pd.Timestamp(fecha_raw).date()

            conductor_raw = str(row.get("conductor", "")).strip() if not pd.isna(row.get("conductor")) else None
            seller_raw = str(row.get("seller", "")).strip() if not pd.isna(row.get("seller")) else None

            if not conductor_raw or not seller_raw:
                errores.append(f"Fila {idx + 2}: faltan conductor o seller")
                continue

            driver_match, driver_score = _mejor_match(conductor_raw, drivers)
            pickup_match, pickup_score = _mejor_match(seller_raw, pickups_list)
            seller_match, seller_score = _mejor_match(seller_raw, sellers)
            suc_match, suc_score = _mejor_match(seller_raw, sucursales_list)

            # Priority: Pickup > Sucursal > Seller (prefer more specific match)
            es_pickup = pickup_match is not None and pickup_score >= max(seller_score, suc_score) and pickup_score >= 0.45
            es_sucursal = not es_pickup and suc_match is not None and suc_score > seller_score and suc_score >= 0.45

            if es_pickup:
                punto = pickup_match
                items.append({
                    "fila": idx + 2,
                    "fecha": str(fecha),
                    "conductor_raw": conductor_raw,
                    "seller_raw": seller_raw,
                    "tipo": "pickup",
                    "driver_id": driver_match.id if driver_match else None,
                    "driver_nombre": driver_match.nombre if driver_match else None,
                    "driver_score": driver_score,
                    "seller_id": None,
                    "seller_nombre": None,
                    "seller_score": 0,
                    "pickup_id": punto.id,
                    "pickup_nombre": punto.nombre,
                    "pickup_score": pickup_score,
                    "sucursal_id": None,
                    "sucursal_nombre": None,
                    "sucursal_score": 0,
                    "tarifa_seller": 0,
                    "tarifa_driver": punto.tarifa_driver or 0,
                })
            elif es_sucursal:
                suc_obj = suc_match
                items.append({
                    "fila": idx + 2,
                    "fecha": str(fecha),
                    "conductor_raw": conductor_raw,
                    "seller_raw": seller_raw,
                    "tipo": "sucursal",
                    "driver_id": driver_match.id if driver_match else None,
                    "driver_nombre": driver_match.nombre if driver_match else None,
                    "driver_score": driver_score,
                    "seller_id": suc_obj.seller_id,
                    "seller_nombre": None,
                    "seller_score": 0,
                    "pickup_id": None,
                    "pickup_nombre": None,
                    "pickup_score": 0,
                    "sucursal_id": suc_obj.id,
                    "sucursal_nombre": suc_obj.nombre,
                    "sucursal_score": suc_score,
                    "tarifa_seller": suc_obj.tarifa_retiro or 0,
                    "tarifa_driver": suc_obj.tarifa_retiro_driver or 0,
                })
            else:
                seller_obj = seller_match
                skip = seller_obj and seller_obj.usa_pickup
                items.append({
                    "fila": idx + 2,
                    "fecha": str(fecha),
                    "conductor_raw": conductor_raw,
                    "seller_raw": seller_raw,
                    "tipo": "seller",
                    "driver_id": driver_match.id if driver_match else None,
                    "driver_nombre": driver_match.nombre if driver_match else None,
                    "driver_score": driver_score,
                    "seller_id": seller_obj.id if seller_obj and not skip else None,
                    "seller_nombre": seller_obj.nombre if seller_obj and not skip else None,
                    "seller_score": seller_score if not skip else 0,
                    "pickup_id": None,
                    "pickup_nombre": None,
                    "pickup_score": 0,
                    "sucursal_id": None,
                    "sucursal_nombre": None,
                    "sucursal_score": 0,
                    "tarifa_seller": (seller_obj.tarifa_retiro or 0) if seller_obj and not skip else 0,
                    "tarifa_driver": (seller_obj.tarifa_retiro_driver or 0) if seller_obj and not skip else 0,
                    "skip_pickup": bool(skip),
                })
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {str(e)}")

    todos_drivers = [{"id": d.id, "nombre": d.nombre} for d in sorted(drivers, key=lambda x: x.nombre)]
    todos_sellers = [{"id": s.id, "nombre": s.nombre} for s in sorted(sellers, key=lambda x: x.nombre) if not s.usa_pickup]
    todos_pickups = [{"id": p.id, "nombre": p.nombre} for p in sorted(pickups_list, key=lambda x: x.nombre)]

    sellers_by_id = {s.id: s.nombre for s in sellers}
    todos_sucursales = [
        {"id": s.id, "nombre": s.nombre, "seller_id": s.seller_id,
         "seller_nombre": sellers_by_id.get(s.seller_id, "?")}
        for s in sorted(sucursales_list, key=lambda x: x.nombre)
    ]

    return {
        "items": items,
        "errores": errores,
        "drivers": todos_drivers,
        "sellers": todos_sellers,
        "pickups": todos_pickups,
        "sucursales": todos_sucursales,
        "archivo": file.filename,
    }


class ItemConfirmarRetiro(BaseModel):
    fila: int
    fecha: str
    conductor_raw: str
    seller_raw: str
    tipo: str  # "seller" | "pickup" | "sucursal"
    driver_id: Optional[int] = None
    seller_id: Optional[int] = None
    pickup_id: Optional[int] = None
    sucursal_id: Optional[int] = None


class ConfirmarRetirosRequest(BaseModel):
    items: List[ItemConfirmarRetiro]
    archivo: Optional[str] = None


@router.post("/importar/confirmar")
def confirmar_retiros(
    body: ConfirmarRetirosRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user=require_permission("retiros:editar"),
):
    """Crea retiros confirmados y guarda aliases de homologación."""
    ingesta_id = str(uuid.uuid4())[:8]
    creados = 0
    creados_pickup = 0
    creados_sucursal = 0

    for item in body.items:
        fecha = pd.to_datetime(item.fecha).date()
        semana = calcular_semana_del_mes(fecha)

        _verificar_semana_abierta(
            db, semana, fecha.month, fecha.year,
            driver_id=item.driver_id, seller_id=item.seller_id, pickup_id=item.pickup_id,
        )

        if item.tipo == "pickup" and item.pickup_id:
            pickup = db.get(Pickup, item.pickup_id)
            driver_obj = db.get(Driver, item.driver_id) if item.driver_id else None
            tarifa_driver_efectiva = pickup.tarifa_driver if pickup else 0
            if driver_obj and driver_obj.tarifa_retiro_fija and driver_obj.tarifa_retiro_fija > 0:
                tarifa_driver_efectiva = driver_obj.tarifa_retiro_fija
            retiro = Retiro(
                fecha=fecha, semana=semana, mes=fecha.month, anio=fecha.year,
                seller_id=None, driver_id=item.driver_id, pickup_id=item.pickup_id,
                tarifa_seller=0,
                tarifa_driver=tarifa_driver_efectiva,
                seller_nombre_raw=item.seller_raw, driver_nombre_raw=item.conductor_raw,
                homologado=item.driver_id is not None,
                ingesta_id=ingesta_id,
            )
            db.add(retiro)
            creados_pickup += 1
        elif item.tipo == "sucursal" and item.sucursal_id:
            suc = db.get(Sucursal, item.sucursal_id)
            driver_obj = db.get(Driver, item.driver_id) if item.driver_id else None
            tarifa_driver_efectiva = suc.tarifa_retiro_driver if suc else 0
            if driver_obj and driver_obj.tarifa_retiro_fija and driver_obj.tarifa_retiro_fija > 0:
                tarifa_driver_efectiva = driver_obj.tarifa_retiro_fija
            retiro = Retiro(
                fecha=fecha, semana=semana, mes=fecha.month, anio=fecha.year,
                seller_id=suc.seller_id if suc else item.seller_id,
                driver_id=item.driver_id,
                sucursal_id=item.sucursal_id,
                tarifa_seller=suc.tarifa_retiro if suc else 0,
                tarifa_driver=tarifa_driver_efectiva,
                seller_nombre_raw=item.seller_raw, driver_nombre_raw=item.conductor_raw,
                homologado=item.driver_id is not None,
                ingesta_id=ingesta_id,
            )
            db.add(retiro)
            creados_sucursal += 1
        elif item.tipo == "seller":
            seller = db.get(Seller, item.seller_id) if item.seller_id else None
            driver_obj = db.get(Driver, item.driver_id) if item.driver_id else None
            tarifa_seller = (seller.tarifa_retiro or 0) if seller else 0
            tarifa_driver = (seller.tarifa_retiro_driver or 0) if seller else 0
            if driver_obj and driver_obj.tarifa_retiro_fija and driver_obj.tarifa_retiro_fija > 0:
                tarifa_driver = driver_obj.tarifa_retiro_fija
            retiro = Retiro(
                fecha=fecha, semana=semana, mes=fecha.month, anio=fecha.year,
                seller_id=item.seller_id, driver_id=item.driver_id,
                tarifa_seller=tarifa_seller, tarifa_driver=tarifa_driver,
                seller_nombre_raw=item.seller_raw, driver_nombre_raw=item.conductor_raw,
                homologado=item.driver_id is not None and item.seller_id is not None,
                ingesta_id=ingesta_id,
            )
            db.add(retiro)
            creados += 1

        # Guardar alias para driver
        if item.driver_id and item.conductor_raw:
            driver = db.get(Driver, item.driver_id)
            if driver:
                alias = item.conductor_raw.strip()
                aliases_lower = [a.lower() for a in (driver.aliases or [])]
                if alias.lower() != driver.nombre.lower() and alias.lower() not in aliases_lower:
                    driver.aliases = list(driver.aliases or []) + [alias]
                    flag_modified(driver, "aliases")

        # Guardar alias para seller
        if item.seller_id and item.seller_raw and item.tipo == "seller":
            seller = db.get(Seller, item.seller_id)
            if seller:
                alias = item.seller_raw.strip()
                aliases_lower = [a.lower() for a in (seller.aliases or [])]
                if alias.lower() != seller.nombre.lower() and alias.lower() not in aliases_lower:
                    seller.aliases = list(seller.aliases or []) + [alias]
                    flag_modified(seller, "aliases")

        # Guardar alias para pickup
        if item.pickup_id and item.seller_raw and item.tipo == "pickup":
            pickup = db.get(Pickup, item.pickup_id)
            if pickup:
                alias = item.seller_raw.strip()
                aliases_lower = [a.lower() for a in (pickup.aliases or [])]
                if alias.lower() != pickup.nombre.lower() and alias.lower() not in aliases_lower:
                    pickup.aliases = list(pickup.aliases or []) + [alias]
                    flag_modified(pickup, "aliases")

        # Guardar alias para sucursal
        if item.sucursal_id and item.seller_raw and item.tipo == "sucursal":
            suc = db.get(Sucursal, item.sucursal_id)
            if suc:
                alias = item.seller_raw.strip()
                aliases_lower = [a.lower() for a in (suc.aliases or [])]
                if alias.lower() != suc.nombre.lower() and alias.lower() not in aliases_lower:
                    suc.aliases = list(suc.aliases or []) + [alias]
                    flag_modified(suc, "aliases")

    periodos_afectados = set()
    for item in body.items:
        fecha = pd.to_datetime(item.fecha).date()
        semana = calcular_semana_del_mes(fecha)
        periodos_afectados.add((semana, fecha.month, fecha.year))
    for s, m, a in periodos_afectados:
        invalidar_snapshots(db, s, m, a)

    db.commit()
    audit(db, "importar_retiros", usuario=current_user, request=request,
          entidad="retiro_batch",
          metadata={"archivo": body.archivo, "creados": creados,
                    "creados_pickup": creados_pickup, "creados_sucursal": creados_sucursal})

    return {"ok": True, "creados": creados, "creados_pickup": creados_pickup, "creados_sucursal": creados_sucursal}


@router.post("/importar")
async def importar_retiros(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=require_permission("retiros:editar"),
):
    """Legacy: importa retiros directamente sin preview."""
    raise HTTPException(status_code=410, detail="Use /importar/preview + /importar/confirmar")


@router.post("", response_model=RetiroOut, status_code=201)
def crear_retiro(data: RetiroCreate, request: Request, db: Session = Depends(get_db), current_user=require_permission("retiros:editar")):
    if not data.seller_id and not data.pickup_id and not data.sucursal_id:
        raise HTTPException(status_code=400, detail="Debe especificar seller, pickup o sucursal")
    if data.seller_id:
        if not db.get(Seller, data.seller_id):
            raise HTTPException(status_code=404, detail="Seller no encontrado")
    if data.pickup_id:
        from app.models import Pickup
        if not db.get(Pickup, data.pickup_id):
            raise HTTPException(status_code=404, detail="Pickup no encontrado")
    driver = db.get(Driver, data.driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    campos = data.model_dump()
    campos["semana"] = calcular_semana_del_mes(data.fecha)
    campos["mes"] = data.fecha.month
    campos["anio"] = data.fecha.year
    _verificar_semana_abierta(
        db, campos["semana"], campos["mes"], campos["anio"],
        driver_id=data.driver_id, seller_id=data.seller_id, pickup_id=data.pickup_id,
    )
    # Si el driver tiene tarifa fija, guardarla en tarifa_driver del retiro (valor histórico)
    if driver.tarifa_retiro_fija and driver.tarifa_retiro_fija > 0:
        campos["tarifa_driver"] = driver.tarifa_retiro_fija
    retiro = Retiro(**campos)
    db.add(retiro)
    invalidar_snapshots(db, campos["semana"], campos["mes"], campos["anio"])
    db.commit()
    db.refresh(retiro)
    audit(db, "crear_retiro", usuario=current_user, request=request, entidad="retiro", entidad_id=retiro.id, metadata={"seller_id": data.seller_id, "pickup_id": data.pickup_id, "driver_id": data.driver_id, "fecha": str(data.fecha)})
    return _enrich_retiro(retiro, db)


@router.put("/{retiro_id}", response_model=RetiroOut)
def editar_retiro(
    retiro_id: int,
    data: RetiroUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user=require_permission("retiros:editar"),
):
    retiro = db.get(Retiro, retiro_id)
    if not retiro:
        raise HTTPException(status_code=404, detail="Retiro no encontrado")

    _verificar_semana_abierta(
        db, retiro.semana, retiro.mes, retiro.anio,
        driver_id=retiro.driver_id, seller_id=retiro.seller_id, pickup_id=retiro.pickup_id,
    )

    antes = {"seller_id": retiro.seller_id, "driver_id": retiro.driver_id,
             "fecha": str(retiro.fecha), "tarifa_seller": retiro.tarifa_seller, "tarifa_driver": retiro.tarifa_driver}

    if data.fecha is not None:
        retiro.fecha = data.fecha
    if data.seller_id is not None:
        if not db.get(Seller, data.seller_id):
            raise HTTPException(status_code=404, detail="Seller no encontrado")
        retiro.seller_id = data.seller_id
    if data.driver_id is not None:
        if not db.get(Driver, data.driver_id):
            raise HTTPException(status_code=404, detail="Driver no encontrado")
        retiro.driver_id = data.driver_id
    if data.tarifa_seller is not None:
        retiro.tarifa_seller = data.tarifa_seller
    if data.tarifa_driver is not None:
        retiro.tarifa_driver = data.tarifa_driver

    invalidar_snapshots(db, retiro.semana, retiro.mes, retiro.anio)
    db.commit()
    db.refresh(retiro)

    despues = {"seller_id": retiro.seller_id, "driver_id": retiro.driver_id,
               "fecha": str(retiro.fecha), "tarifa_seller": retiro.tarifa_seller, "tarifa_driver": retiro.tarifa_driver}
    cambios = {k: {"antes": antes[k], "despues": despues[k]} for k in antes if antes[k] != despues[k]}
    if cambios:
        audit(db, "editar_retiro", usuario=current_user, request=request,
              entidad="retiro", entidad_id=retiro_id, cambios=cambios)
    return _enrich_retiro(retiro, db)


@router.delete("/batch")
def eliminar_retiros_batch(
    ids: List[int],
    request: Request,
    db: Session = Depends(get_db),
    current_user=require_permission("retiros:editar"),
):
    """Elimina múltiples retiros por lista de IDs."""
    if not ids:
        raise HTTPException(status_code=400, detail="Lista de IDs vacía")

    retiros = db.query(Retiro).filter(Retiro.id.in_(ids)).all()
    if not retiros:
        raise HTTPException(status_code=404, detail="No se encontraron retiros")

    cerrados = [
        r for r in retiros
        if _check_cerrado(db, r.semana, r.mes, r.anio, r.driver_id, r.seller_id, r.pickup_id)
    ]
    if cerrados:
        raise HTTPException(
            status_code=409,
            detail=f"{len(cerrados)} retiro(s) pertenecen a semanas ya cerradas y no pueden eliminarse.",
        )

    periodos = set()
    for r in retiros:
        periodos.add((r.semana, r.mes, r.anio))
        db.delete(r)

    for s, m, a in periodos:
        invalidar_snapshots(db, s, m, a)

    db.commit()
    audit(db, "eliminar_retiros_batch", usuario=current_user, request=request,
          entidad="retiro_batch", metadata={"ids": ids, "eliminados": len(retiros)})
    return {"message": f"{len(retiros)} retiros eliminados", "eliminados": len(retiros)}


@router.delete("/{retiro_id}")
def eliminar_retiro(retiro_id: int, request: Request, db: Session = Depends(get_db), current_user=require_permission("retiros:editar")):
    retiro = db.get(Retiro, retiro_id)
    if not retiro:
        raise HTTPException(status_code=404, detail="Retiro no encontrado")
    _verificar_semana_abierta(
        db, retiro.semana, retiro.mes, retiro.anio,
        driver_id=retiro.driver_id, seller_id=retiro.seller_id, pickup_id=retiro.pickup_id,
    )
    meta = {"seller_id": retiro.seller_id, "driver_id": retiro.driver_id, "fecha": str(retiro.fecha)}
    semana, mes, anio = retiro.semana, retiro.mes, retiro.anio
    db.delete(retiro)
    invalidar_snapshots(db, semana, mes, anio)
    db.commit()
    audit(db, "eliminar_retiro", usuario=current_user, request=request, entidad="retiro", entidad_id=retiro_id, metadata=meta)
    return {"message": "Retiro eliminado"}
