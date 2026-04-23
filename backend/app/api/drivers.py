from typing import Optional, List
import io

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from sqlalchemy import func as sqlfunc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from app.database import get_db
from app.auth import require_admin, require_admin_or_administracion, require_permission, require_driver, get_current_user, hash_password
from app.models import Driver, RolEnum, Retiro, PagoSemanaDriver, EstadoPagoEnum
from app.schemas import DriverCreate, DriverUpdate, DriverOut, AcuerdoAceptarRequest, ContratoTrabajoAceptarRequest
from app.services.audit import registrar as audit
from app.services.audit import diff_campos

router = APIRouter(prefix="/drivers", tags=["Drivers"])

CURRENT_CONTRATO_TRABAJO_VERSION = "1.0"


def _normalizar_rut(valor: str) -> str:
    """Quita puntos, guiones y espacios del RUT; normaliza DV a mayúscula.
    Ej: '12.733.783-8', '127337838', '12733783-8' → '127337838'."""
    if not valor:
        return ""
    limpio = "".join(ch for ch in str(valor) if ch.isalnum()).upper()
    return limpio


def _normalizar_nombre(valor: str) -> str:
    """Normaliza nombre para comparación: quita tildes/diacríticos, colapsa
    espacios y baja a minúsculas."""
    import unicodedata
    if not valor:
        return ""
    s = unicodedata.normalize("NFKD", str(valor))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = " ".join(s.split()).lower()
    return s


def _enrich_driver(d: Driver, db: Session) -> dict:
    from app.api.auth import CURRENT_ACUERDO_VERSION
    data = {c.name: getattr(d, c.name) for c in d.__table__.columns}
    data.pop("acuerdo_firma", None)
    data.pop("carnet_frontal", None)
    data.pop("carnet_trasero", None)
    # Recalcular acuerdo_aceptado con chequeo de versión vigente
    data["acuerdo_aceptado"] = bool(d.contratado) or bool(
        d.acuerdo_aceptado and d.acuerdo_version == CURRENT_ACUERDO_VERSION
    )
    data["aliases"] = d.aliases or []
    if d.jefe_flota_id:
        jefe = db.get(Driver, d.jefe_flota_id)
        data["jefe_flota_nombre"] = jefe.nombre if jefe else None
    else:
        data["jefe_flota_nombre"] = None
    data["subordinados_count"] = db.query(Driver).filter(
        Driver.jefe_flota_id == d.id, Driver.activo == True
    ).count()
    if d.trabajador_id:
        from app.models import Trabajador
        trab = db.get(Trabajador, d.trabajador_id)
        data["trabajador_nombre"] = trab.nombre if trab else None
    else:
        data["trabajador_nombre"] = None
    return data


@router.get("", response_model=List[DriverOut])
def listar_drivers(
    activo: Optional[bool] = None,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    query = db.query(Driver)
    if activo is not None:
        query = query.filter(Driver.activo == activo)
    if q:
        query = query.filter(Driver.nombre.ilike(f"%{q}%"))
    drivers = query.order_by(Driver.nombre).all()
    return [_enrich_driver(d, db) for d in drivers]


@router.get("/plantilla/homologacion/descargar")
def descargar_plantilla_homologacion(db: Session = Depends(get_db)):
    """Genera plantilla Excel para homologación de nombres de drivers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Homologación"

    columnas = ["Nombre Driver Raw", "Nombre Driver"]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    for i, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = border

    drivers = db.query(Driver).filter(Driver.activo == True).order_by(Driver.nombre).all()
    alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    ri = 2
    for d in drivers:
        for alias in (d.aliases or []):
            ws.cell(row=ri, column=1, value=alias).font = Font(size=10)
            ws.cell(row=ri, column=2, value=d.nombre).font = Font(size=10)
            for ci in (1, 2):
                ws.cell(row=ri, column=ci).border = border
                if ri % 2 == 0:
                    ws.cell(row=ri, column=ci).fill = alt
            ri += 1

    if ri == 2:
        ejemplos = [
            ["Carlos Pérez (auto-create)", "Carlos"],
            ["Carlos P.", "Carlos"],
            ["AUGUSTO", "Augusto"],
            ["Jorge Diaz", "Jorge"],
        ]
        for fila in ejemplos:
            for ci, val in enumerate(fila, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(size=10)
                c.border = border
                if ri % 2 == 0:
                    c.fill = alt
            ri += 1

    anchos = [36, 30]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.auto_filter.ref = "A1:B1"
    ws.freeze_panes = "A2"

    wsi = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("Plantilla de Homologación de Drivers", ""),
        ("", ""),
        ("Campo", "Descripción"),
        ("Nombre Driver Raw", "Nombre tal como llega desde el software de gestión (ej: 'Carlos Pérez (auto-create)', 'AUGUSTO')."),
        ("Nombre Driver", "Nombre oficial del conductor en el sistema. Si no existe, se creará automáticamente."),
        ("", ""),
        ("Nota", "Cada fila mapea un nombre raw a un driver. Un driver puede tener múltiples nombres raw."),
    ]
    for ri, (a, b) in enumerate(instrucciones, 1):
        ca = wsi.cell(row=ri, column=1, value=a)
        cb = wsi.cell(row=ri, column=2, value=b)
        if ri == 1:
            ca.font = Font(bold=True, size=14, color="1A365D")
        elif ri == 3:
            ca.font = Font(bold=True, size=11)
            cb.font = Font(bold=True, size=11)
        else:
            ca.font = Font(bold=True, size=10) if a else Font(size=10)
            cb.font = Font(size=10, color="4A5568")
    wsi.column_dimensions["A"].width = 22
    wsi.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_homologacion_drivers.xlsx"})


@router.get("/plantilla/tarifas/descargar")
def descargar_plantilla_tarifas(db: Session = Depends(get_db)):
    """Genera plantilla Excel para tarifas de drivers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tarifas"

    columnas = ["Nombre Driver", "Tarifa ECOURIER", "Tarifa OVIEDO", "Tarifa TERCERIZADO"]

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    for i, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = border

    drivers = db.query(Driver).filter(Driver.activo == True).order_by(Driver.nombre).all()
    alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    for ri, d in enumerate(drivers, 2):
        vals = [d.nombre, d.tarifa_ecourier, d.tarifa_oviedo, d.tarifa_tercerizado]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=10)
            c.border = border
            if ri % 2 == 0:
                c.fill = alt

    if not drivers:
        ejemplos = [
            ["Carlos", 1700, 1800, 1500],
            ["Augusto", 1900, 1800, 1500],
            ["Jorge", 2000, 1800, 1500],
        ]
        for ri, fila in enumerate(ejemplos, 2):
            for ci, val in enumerate(fila, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(size=10)
                c.border = border
                if ri % 2 == 0:
                    c.fill = alt

    anchos = [28, 18, 18, 22]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.auto_filter.ref = "A1:D1"
    ws.freeze_panes = "A2"

    wsi = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("Plantilla de Tarifas de Drivers", ""),
        ("", ""),
        ("Campo", "Descripción"),
        ("Nombre Driver", "Nombre oficial del conductor (obligatorio, único). Si no existe, se creará."),
        ("Tarifa ECOURIER", "Monto en CLP por envío de seller tipo ECOURIER."),
        ("Tarifa OVIEDO", "Monto en CLP por envío de seller tipo OVIEDO."),
        ("Tarifa TERCERIZADO", "Monto en CLP por envío de seller tipo TERCERIZADO."),
    ]
    for ri, (a, b) in enumerate(instrucciones, 1):
        ca = wsi.cell(row=ri, column=1, value=a)
        cb = wsi.cell(row=ri, column=2, value=b)
        if ri == 1:
            ca.font = Font(bold=True, size=14, color="1A365D")
        elif ri == 3:
            ca.font = Font(bold=True, size=11)
            cb.font = Font(bold=True, size=11)
        else:
            ca.font = Font(bold=True, size=10) if a else Font(size=10)
            cb.font = Font(size=10, color="4A5568")
    wsi.column_dimensions["A"].width = 22
    wsi.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_tarifas_drivers.xlsx"})


@router.post("/importar/homologacion")
async def importar_homologacion(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("drivers:editar")),
):
    """
    Importa plantilla de homologación.
    Columnas: Nombre Driver Raw, Nombre Driver.
    Agrega cada raw_name como alias del driver. Si el driver no existe, lo crea.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")

    content = await file.read()
    df = pd.read_excel(io.BytesIO(content))

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "raw" in cl:
            col_map[col] = "raw_name"
        elif "driver" in cl or "conductor" in cl or "nombre" in cl:
            if "raw_name" not in col_map.values():
                col_map[col] = "raw_name"
            else:
                col_map[col] = "driver_name"
    df = df.rename(columns=col_map)

    if "raw_name" not in df.columns or "driver_name" not in df.columns:
        cols = list(df.columns)
        if len(cols) >= 2:
            df = df.rename(columns={cols[0]: "raw_name", cols[1]: "driver_name"})
        else:
            raise HTTPException(status_code=400, detail="Se esperan al menos 2 columnas: Nombre Driver Raw, Nombre Driver")

    aliases_added = 0
    drivers_created = 0
    errores = []

    for idx, row in df.iterrows():
        try:
            raw_name = str(row.get("raw_name", "")).strip()
            driver_name = str(row.get("driver_name", "")).strip()

            if not raw_name or raw_name == "nan" or not driver_name or driver_name == "nan":
                continue

            driver = db.query(Driver).filter(
                sqlfunc.lower(Driver.nombre) == driver_name.lower()
            ).first()
            if not driver:
                driver = Driver(nombre=driver_name)
                db.add(driver)
                db.flush()
                drivers_created += 1

            current_aliases = list(driver.aliases or [])
            already = any(a.lower() == raw_name.lower() for a in current_aliases)
            if not already:
                current_aliases.append(raw_name)
                driver.aliases = current_aliases
                flag_modified(driver, "aliases")
                aliases_added += 1

        except Exception as e:
            errores.append(f"Fila {idx + 2}: {str(e)}")

    db.commit()
    audit(
        db, "importar_homologacion_driver",
        usuario=current_user, request=request,
        entidad="driver",
        metadata={"archivo": file.filename, "aliases_agregados": aliases_added,
                   "drivers_creados": drivers_created},
    )
    return {
        "aliases_agregados": aliases_added,
        "drivers_creados": drivers_created,
        "errores": errores,
    }


@router.post("/importar/tarifas")
async def importar_tarifas(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("drivers:editar")),
):
    """
    Importa plantilla de tarifas.
    Columnas: Nombre Driver, Tarifa ECOURIER, Tarifa OVIEDO, Tarifa TERCERIZADO.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")

    content = await file.read()
    df = pd.read_excel(io.BytesIO(content))

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "ecourier" in cl:
            col_map[col] = "tarifa_ecourier"
        elif "oviedo" in cl:
            col_map[col] = "tarifa_oviedo"
        elif "tercerizado" in cl:
            col_map[col] = "tarifa_tercerizado"
        elif ("nombre" in cl or "driver" in cl or "conductor" in cl) and "tarifa" not in cl:
            col_map[col] = "nombre"
    df = df.rename(columns=col_map)

    if "nombre" not in df.columns:
        cols = [c for c in df.columns if c not in col_map.values()]
        if cols:
            df = df.rename(columns={cols[0]: "nombre"})
        else:
            raise HTTPException(status_code=400, detail="No se encontró columna de nombre del driver")

    creados = 0
    actualizados = 0
    errores = []
    seen_names = set()

    for idx, row in df.iterrows():
        try:
            nombre = str(row.get("nombre", "")).strip()
            if not nombre or nombre == "nan":
                continue
            if nombre in seen_names:
                continue
            seen_names.add(nombre)

            t_ec = int(row.get("tarifa_ecourier", 1700)) if not pd.isna(row.get("tarifa_ecourier", 0)) else 1700
            t_ov = int(row.get("tarifa_oviedo", 1800)) if not pd.isna(row.get("tarifa_oviedo", 0)) else 1800
            t_te = int(row.get("tarifa_tercerizado", 1500)) if not pd.isna(row.get("tarifa_tercerizado", 0)) else 1500

            existing = db.query(Driver).filter(
                sqlfunc.lower(Driver.nombre) == nombre.lower()
            ).first()
            if existing:
                existing.tarifa_ecourier = t_ec
                existing.tarifa_oviedo = t_ov
                existing.tarifa_tercerizado = t_te
                actualizados += 1
            else:
                driver = Driver(
                    nombre=nombre,
                    tarifa_ecourier=t_ec, tarifa_oviedo=t_ov, tarifa_tercerizado=t_te,
                )
                db.add(driver)
                creados += 1
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {str(e)}")

    db.commit()
    audit(
        db, "importar_tarifas_driver",
        usuario=current_user, request=request,
        entidad="driver",
        metadata={"archivo": file.filename, "creados": creados,
                   "actualizados": actualizados},
    )
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


@router.get("/plantilla/bancaria/descargar")
def descargar_plantilla_bancaria(db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    """Plantilla Excel con drivers para completar datos bancarios."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Bancarios"

    columnas = ["Conductor", "RUT", "Banco", "Tipo Cuenta", "N° Cuenta"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    for i, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = border

    drivers = db.query(Driver).filter(Driver.activo == True).order_by(Driver.nombre).all()
    alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    for ri, d in enumerate(drivers, 2):
        vals = [d.nombre, d.rut or "", d.banco or "", d.tipo_cuenta or "", d.numero_cuenta or ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=10, bold=(ci == 1))
            c.border = border
            if ri % 2 == 0:
                c.fill = alt

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 22
    ws.auto_filter.ref = "A1:E1"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=plantilla_bancaria_drivers.xlsx"})


@router.post("/importar/bancaria")
async def importar_bancaria(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_permission("drivers:editar")),
):
    """Importa datos bancarios de drivers desde Excel."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")

    content = await file.read()
    df = pd.read_excel(io.BytesIO(content))

    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if "conductor" in cl or "driver" in cl or "nombre" in cl:
            col_map[col] = "nombre"
        elif "rut" in cl:
            col_map[col] = "rut"
        elif "banco" in cl:
            col_map[col] = "banco"
        elif "tipo" in cl and "cuenta" in cl:
            col_map[col] = "tipo_cuenta"
        elif ("cuenta" in cl or "numero" in cl) and "tipo" not in cl:
            col_map[col] = "numero_cuenta"
    df = df.rename(columns=col_map)

    actualizados = 0
    errores = []

    for idx, row in df.iterrows():
        try:
            nombre = str(row.get("nombre", "")).strip()
            if not nombre or nombre == "nan":
                continue

            driver = db.query(Driver).filter(
                sqlfunc.lower(Driver.nombre) == nombre.lower()
            ).first()
            if not driver:
                errores.append(f"Fila {idx + 2}: driver '{nombre}' no encontrado")
                continue

            for field in ["rut", "banco", "tipo_cuenta", "numero_cuenta"]:
                val = str(row.get(field, "")).strip() if not pd.isna(row.get(field, "")) else None
                if val and val != "nan":
                    setattr(driver, field, val)

            actualizados += 1
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {str(e)}")

    db.commit()
    audit(
        db, "importar_bancaria_driver",
        usuario=current_user, request=request,
        entidad="driver",
        metadata={"archivo": file.filename, "actualizados": actualizados},
    )
    return {"actualizados": actualizados, "errores": errores}


@router.get("/{driver_id}", response_model=DriverOut)
def obtener_driver(driver_id: int, db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    driver = db.query(Driver).get(driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    return _enrich_driver(driver, db)


@router.post("", response_model=DriverOut, status_code=201)
def crear_driver(data: DriverCreate, request: Request, db: Session = Depends(get_db), current_user=Depends(require_permission("drivers:editar"))):
    existing = db.query(Driver).filter(Driver.nombre == data.nombre).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un driver con ese nombre")
    driver = Driver(**data.model_dump(exclude={"password"}))
    if data.password:
        driver.password_hash = hash_password(data.password)
    db.add(driver)
    db.commit()
    db.refresh(driver)
    audit(
        db, "crear_driver",
        usuario=current_user, request=request,
        entidad="driver", entidad_id=driver.id,
        metadata={"nombre": driver.nombre, "contratado": driver.contratado,
                   "tarifa_ecourier": driver.tarifa_ecourier,
                   "tarifa_oviedo": driver.tarifa_oviedo,
                   "tarifa_tercerizado": driver.tarifa_tercerizado},
    )
    return _enrich_driver(driver, db)


@router.put("/{driver_id}", response_model=DriverOut)
def actualizar_driver(
    driver_id: int, data: DriverUpdate, request: Request,
    db: Session = Depends(get_db), current_user=Depends(require_permission("drivers:editar")),
):
    driver = db.query(Driver).get(driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    campos_audit = [
        "nombre", "tarifa_ecourier", "tarifa_oviedo", "tarifa_tercerizado",
        "tarifa_retiro_fija", "aliases", "contratado", "jefe_flota_id",
        "rut", "banco", "tipo_cuenta", "numero_cuenta", "email", "trabajador_id",
    ]
    antes = {c: getattr(driver, c, None) for c in campos_audit}
    update_data = data.model_dump(exclude_unset=True, exclude={"password"})
    for nullable_unique in ("email", "rut"):
        if nullable_unique in update_data and update_data[nullable_unique] == "":
            update_data[nullable_unique] = None
    for key, value in update_data.items():
        setattr(driver, key, value)
    if "trabajador_id" in update_data:
        driver.contratado = update_data["trabajador_id"] is not None
    if data.password:
        driver.password_hash = hash_password(data.password)
    db.flush()
    despues = {c: getattr(driver, c, None) for c in campos_audit}
    cambios = diff_campos(antes, despues, campos_audit)

    # Si cambió tarifa_retiro_fija, actualizar Retiro.tarifa_driver en semanas NO cerradas
    nueva_tarifa_fija = despues.get("tarifa_retiro_fija")
    vieja_tarifa_fija = antes.get("tarifa_retiro_fija")
    if nueva_tarifa_fija != vieja_tarifa_fija and nueva_tarifa_fija is not None and nueva_tarifa_fija > 0:
        # Obtener semanas cerradas para este driver
        semanas_cerradas = db.query(
            PagoSemanaDriver.semana, PagoSemanaDriver.mes, PagoSemanaDriver.anio,
        ).filter(
            PagoSemanaDriver.driver_id == driver_id,
            PagoSemanaDriver.estado == EstadoPagoEnum.PAGADO.value,
        ).all()
        semanas_cerradas_set = {(r.semana, r.mes, r.anio) for r in semanas_cerradas}

        # Actualizar solo retiros de semanas abiertas
        retiros_abiertos = db.query(Retiro).filter(
            Retiro.driver_id == driver_id,
        ).all()
        for retiro in retiros_abiertos:
            if (retiro.semana, retiro.mes, retiro.anio) not in semanas_cerradas_set:
                retiro.tarifa_driver = nueva_tarifa_fija

    if cambios:
        audit(
            db, "editar_driver",
            usuario=current_user, request=request,
            entidad="driver", entidad_id=driver.id,
            cambios=cambios,
            metadata={"nombre": driver.nombre},
        )
    db.commit()
    db.refresh(driver)
    return _enrich_driver(driver, db)


@router.delete("/{driver_id}")
def eliminar_driver(driver_id: int, request: Request, db: Session = Depends(get_db), current_user=Depends(require_permission("drivers:editar"))):
    driver = db.query(Driver).get(driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    driver.activo = False
    db.commit()
    audit(
        db, "eliminar_driver",
        usuario=current_user, request=request,
        entidad="driver", entidad_id=driver.id,
        metadata={"nombre": driver.nombre},
    )
    return {"message": "Driver desactivado"}


@router.patch("/{driver_id}/activar")
def activar_driver(driver_id: int, request: Request, db: Session = Depends(get_db), current_user=Depends(require_permission("drivers:editar"))):
    driver = db.query(Driver).get(driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    if driver.activo:
        raise HTTPException(status_code=400, detail="El driver ya está activo")
    driver.activo = True
    db.commit()
    audit(
        db, "activar_driver",
        usuario=current_user, request=request,
        entidad="driver", entidad_id=driver.id,
        metadata={"nombre": driver.nombre},
    )
    return {"message": "Driver activado"}


@router.get("/mi-flota/info")
def obtener_mi_flota(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Devuelve info de flota para el driver logueado."""
    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")
    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    subordinados = db.query(Driver).filter(
        Driver.jefe_flota_id == driver.id, Driver.activo == True
    ).order_by(Driver.nombre).all()
    return {
        "es_jefe_flota": len(subordinados) > 0,
        "subordinados": [{"id": s.id, "nombre": s.nombre} for s in subordinados],
    }


@router.post("/me/acuerdo")
def aceptar_acuerdo(
    body: AcuerdoAceptarRequest,
    request: Request,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Registra la aceptación digital del acuerdo de colaboración."""
    from datetime import datetime, timezone
    from app.auth import create_access_token
    from app.schemas import TokenResponse
    from app.api.auth import CURRENT_ACUERDO_VERSION

    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")

    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    if driver.contratado:
        raise HTTPException(status_code=400, detail="Los conductores contratados no requieren este acuerdo")

    ip = request.headers.get("X-Forwarded-For", "")
    if ip:
        ip = ip.split(",")[0].strip()
    else:
        ip = request.client.host if request.client else "unknown"

    driver.nombre_completo = body.nombre_completo
    driver.rut = body.rut
    driver.acuerdo_aceptado = True
    driver.acuerdo_version = CURRENT_ACUERDO_VERSION
    driver.acuerdo_fecha = datetime.now(timezone.utc)
    driver.acuerdo_ip = ip
    driver.acuerdo_firma = body.firma_base64
    driver.carnet_frontal = body.carnet_frontal
    driver.carnet_trasero = body.carnet_trasero

    audit(
        db, "aceptar_acuerdo",
        usuario=user, request=request,
        entidad="driver", entidad_id=driver.id,
        metadata={"version": CURRENT_ACUERDO_VERSION, "ip": ip, "rut": body.rut, "nombre_completo": body.nombre_completo},
    )
    db.commit()
    db.refresh(driver)

    token = create_access_token({"sub": str(driver.id), "rol": RolEnum.DRIVER})
    es_jefe = db.query(Driver).filter(Driver.jefe_flota_id == driver.id, Driver.activo == True).count() > 0
    return TokenResponse(
        access_token=token,
        rol=RolEnum.DRIVER,
        nombre=driver.nombre,
        entidad_id=driver.id,
        acuerdo_aceptado=True,
        contratado=driver.contratado,
        es_jefe=es_jefe,
    )


@router.get("/me/acuerdo-info")
def mi_acuerdo_info(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Devuelve los detalles del acuerdo aceptado por el driver logueado."""
    from app.api.auth import CURRENT_ACUERDO_VERSION

    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")
    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    return {
        "nombre": driver.nombre,
        "nombre_completo": driver.nombre_completo,
        "rut": driver.rut,
        "contratado": driver.contratado,
        "acuerdo_aceptado": bool(driver.contratado) or bool(driver.acuerdo_aceptado and driver.acuerdo_version == CURRENT_ACUERDO_VERSION),
        "acuerdo_version": driver.acuerdo_version,
        "acuerdo_fecha": driver.acuerdo_fecha.isoformat() if driver.acuerdo_fecha else None,
        "acuerdo_ip": driver.acuerdo_ip,
        "acuerdo_firma": driver.acuerdo_firma,
        "carnet_frontal": driver.carnet_frontal,
        "carnet_trasero": driver.carnet_trasero,
        "version_actual": CURRENT_ACUERDO_VERSION,
        "tarifas": _tarifas_driver(driver),
    }


@router.get("/me/acuerdo-pdf")
def descargar_acuerdo_pdf(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Genera y descarga el PDF del acuerdo firmado por el driver."""
    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")
    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    if not driver.acuerdo_aceptado:
        raise HTTPException(status_code=404, detail="No hay acuerdo firmado registrado")

    pdf_bytes = _generar_pdf_acuerdo(driver)
    nombre = (driver.nombre or "driver").replace(" ", "_").lower()
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="acuerdo_ecourier_{nombre}.pdf"'},
    )


def _generar_pdf_acuerdo(driver) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import base64, re

    try:
        import font_roboto, os as _os
        _fdir = _os.path.join(_os.path.dirname(font_roboto.__file__), "files")
        pdfmetrics.registerFont(TTFont("Roboto", _os.path.join(_fdir, "Roboto-Regular.ttf")))
        pdfmetrics.registerFont(TTFont("Roboto-Bold", _os.path.join(_fdir, "Roboto-Bold.ttf")))
        FONT, FONTB = "Roboto", "Roboto-Bold"
    except Exception:
        FONT, FONTB = "Helvetica", "Helvetica-Bold"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=2.5*cm, rightMargin=2.5*cm,
        topMargin=2.5*cm, bottomMargin=2.5*cm,
    )

    AZUL = colors.HexColor("#003c72")
    GRIS = colors.HexColor("#6b7280")
    CLARO = colors.HexColor("#f3f4f6")

    styles = getSampleStyleSheet()
    sNormal = ParagraphStyle("n", fontName=FONT, fontSize=9, leading=14, alignment=TA_JUSTIFY, textColor=colors.HexColor("#111827"))
    sTitle  = ParagraphStyle("t", fontName=FONTB, fontSize=16, leading=20, alignment=TA_CENTER, textColor=AZUL, spaceAfter=4)
    sSub    = ParagraphStyle("s", fontName=FONT, fontSize=10, leading=14, alignment=TA_CENTER, textColor=GRIS, spaceAfter=12)
    sClaus  = ParagraphStyle("c", fontName=FONTB, fontSize=9.5, leading=13, textColor=AZUL, spaceBefore=10, spaceAfter=3)
    sMeta   = ParagraphStyle("m", fontName=FONT, fontSize=8.5, leading=13, textColor=GRIS)
    sLegal  = ParagraphStyle("l", fontName=FONT, fontSize=7.5, leading=11, alignment=TA_CENTER, textColor=GRIS)

    fmtCLP = lambda n: f"${n:,.0f}".replace(",", ".")

    clausulas = [
        ("1", "Naturaleza del Acuerdo",
         "Las partes declaran expresamente que el presente acuerdo tiene carácter <b>estrictamente civil y comercial</b>, rigiéndose por el Código Civil y demás normativa aplicable. No existe vínculo de subordinación ni dependencia, ni relación laboral en los términos del Código del Trabajo. El Prestador ejecuta sus servicios de manera autónoma, por cuenta propia, bajo su exclusivo riesgo y responsabilidad, y <b>no forma parte de la estructura organizacional</b> de Ecourier."),
        ("2", "Objeto de los Servicios",
         "El presente acuerdo tiene por objeto la prestación de servicios de retiro, transporte y entrega de mercancías por parte del Prestador, utilizando sus propios medios materiales (vehículo, combustible, dispositivos tecnológicos, etc.). Cada servicio constituye una prestación independiente que debe ser previamente aceptada por el Prestador de forma libre y voluntaria.<br/><br/>Este Acuerdo es de <b>duración indefinida</b> y puede ser terminado en cualquier momento por cualquiera de las partes conforme a la cláusula de Terminación."),
        ("3", "Autonomía, Libertad Operativa y Zona Geográfica",
         "El Prestador gozará de plena autonomía en la organización y ejecución de sus servicios, determinando libremente su disponibilidad, horarios y forma de prestación. Podrá <b>aceptar o rechazar cualquier servicio sin expresión de causa</b>. No existe cláusula de exclusividad; el Prestador puede prestar servicios a terceros sin restricción.<br/><br/>La sugerencia de zonas, rutas o áreas de operación tendrá carácter meramente referencial. El Prestador es responsable de evaluar las condiciones de seguridad, distancia y conveniencia de cada servicio."),
        ("4", "Uso de la Plataforma Digital — Acceso Personal e Intransferible",
         "La plataforma digital de Ecourier es una herramienta tecnológica para facilitar la coordinación y trazabilidad de los servicios. Su utilización <b>no implica el ejercicio de facultades de dirección, control laboral o supervisión jerárquica</b>.<br/><br/>Las credenciales de acceso son <b>estrictamente personales e intransferibles</b>. Queda prohibido cederlas, compartirlas o permitir su uso a terceros. El incumplimiento es causal de término inmediato."),
        ("5", "Coordinación Operativa",
         "Las comunicaciones tendrán carácter estrictamente colaborativo y naturaleza referencial. Se valora que el Prestador informe con anticipación cuando no esté disponible: ausencias de un día con 12 horas; varios días con 15 días; en períodos de alta demanda con <b>30 días</b>. La falta de aviso no genera penalización, pero puede afectar la asignación futura."),
        ("6", "Estándares de Calidad del Servicio",
         "El Prestador ejecutará los servicios con diligencia: manipular la carga con cuidado y verificarla al recibirla; no abrir ni alterar empaques; mantener trato respetuoso con los destinatarios; respetar normas de tránsito; usar canales oficiales de comunicación; y abstenerse de negociaciones paralelas con clientes.<br/><br/>El incumplimiento grave o reiterado puede ser causal de término de la colaboración."),
        ("7", "Régimen de Incumplimientos",
         "<b>a) Leves:</b> No afectan sustancialmente la ejecución del servicio. Generan amonestación escrita.<br/><b>b) Graves:</b> Afectan la ejecución del servicio, la experiencia del cliente o la integridad de la carga. Obligan al Prestador a indemnizar el daño acreditado.<br/><b>c) Reiterados:</b> Dos o más en 30 días. Constituyen causal de término.<br/><br/>Previo a medidas económicas o término, Ecourier notificará al Prestador y otorgará 5 días hábiles para descargos."),
        ("8", "Responsabilidad y Riesgo",
         "El Prestador responde únicamente por daños consecuencia directa de su dolo o culpa grave debidamente acreditados. No responderá por caso fortuito, fuerza mayor o actos de terceros.<br/><br/>La responsabilidad se limita al daño directo con tope en el valor del servicio o carga involucrada, salvo dolo. <b>Ecourier no será responsable por los actos ejecutados por el Prestador en el ejercicio autónomo de su actividad.</b>"),
        ("9", "Seguridad Personal y Protocolo de Robo",
         "El Prestador prioriza en todo momento su <b>integridad física</b> por sobre la carga. Se recomienda el uso de calzado de seguridad y chaleco reflectante.<br/><br/><b>En caso de asalto:</b> no oponer resistencia; comunicar de inmediato a operaciones; realizar denuncia en Carabineros dentro de 24 horas y enviar copia a Ecourier."),
        ("10", "Responsabilidad por Infracciones de Tránsito",
         "Toda multa, infracción o sanción de tránsito es de <b>exclusiva responsabilidad del Prestador</b>, incluyendo las derivadas del estado del vehículo, documentación vencida o conducta al volante. Ecourier no asume ninguna responsabilidad en esta materia."),
        ("11", "Tarifas y Condiciones Económicas",
         "Los servicios serán remunerados conforme a las tarifas vigentes al momento de la prestación. Ecourier podrá modificarlas con <b>30 días de aviso previo</b>; ante cualquier modificación se generará una nueva versión de este Acuerdo.<br/><br/>No existe remuneración fija ni garantía de ingresos mínimos. El Prestador deberá emitir <b>boleta de honorarios o factura</b> como requisito para el pago."),
        ("12", "Protección de Datos Personales",
         "Conforme a la <b>Ley N° 19.628</b>, el Prestador usará los datos personales de destinatarios únicamente para fines de la entrega. El Prestador autoriza a Ecourier para el tratamiento de sus datos personales con la finalidad de gestionar la relación contractual."),
        ("13", "Trazabilidad y Prueba",
         "Las partes acuerdan que toda información relativa a la ejecución de los servicios, incluyendo registros en la plataforma y comunicaciones electrónicas, constituirá <b>medio válido de prueba</b> para determinar el cumplimiento de las obligaciones."),
        ("14", "Confidencialidad",
         "El Prestador mantiene estricta confidencialidad sobre datos de clientes, rutas, tarifas, procesos operativos y estrategias comerciales durante la colaboración y por <b>24 meses después de su término</b>. El incumplimiento puede derivar en acciones legales."),
        ("15", "Terminación",
         "El presente acuerdo podrá ser terminado por cualquiera de las partes en cualquier momento, sin expresión de causa y <b>sin derecho a indemnización</b>, mediante comunicación por correo electrónico o canal oficial.<br/><br/>Son causales de término inmediato: incumplimientos graves o reiterados; faltas graves de conducta con clientes; pérdida de carga por negligencia; falsificación de información; cesión de credenciales a terceros."),
        ("16", "Ley Aplicable y Resolución de Disputas",
         "Este Acuerdo se rige por las leyes de la República de Chile. Las controversias serán sometidas a la jurisdicción de los <b>Tribunales Ordinarios de Justicia de Santiago</b>, renunciando las partes a cualquier otro fuero."),
        ("17", "Modificaciones al Acuerdo",
         "Ecourier podrá modificar los términos de este Acuerdo con <b>30 días de aviso previo</b>. Cualquier modificación generará una nueva versión del Acuerdo que el Prestador deberá aceptar expresamente para continuar prestando servicios."),
        ("18", "Aceptación Digital",
         "Este Acuerdo se acepta de forma electrónica a través de la plataforma de Ecourier, en conformidad con la <b>Ley N° 19.799 sobre Documentos Electrónicos, Firma Electrónica y Servicios de Certificación</b>. La firma digital tiene plena validez legal."),
    ]

    story = []

    # Encabezado
    story.append(Paragraph("ACUERDO DE PRESTACIÓN DE SERVICIOS INDEPENDIENTES", sTitle))
    story.append(Paragraph(f"LOGÍSTICA Y TRANSPORTE E-COURIER SpA · RUT 77.512.163-7 · Moneda 1137 of. 56, Santiago", sSub))
    story.append(HRFlowable(width="100%", thickness=1, color=AZUL))
    story.append(Spacer(1, 10))

    story.append(Paragraph(
        f"En Santiago, entre <b>LOGÍSTICA Y TRANSPORTE E-COURIER SpA</b>, RUT N° 77.512.163-7, con domicilio en "
        f"Moneda N°1137, Oficina 56, comuna de Santiago, en adelante \"Ecourier\"; y <b>{driver.nombre}</b>, "
        f"RUT {driver.rut or '—'}, en adelante el \"Prestador\", se ha convenido el siguiente Acuerdo "
        f"de Prestación de Servicios Independientes (Versión {driver.acuerdo_version}).",
        sNormal,
    ))
    story.append(Spacer(1, 8))

    for num, titulo, texto in clausulas:
        story.append(Paragraph(f"Cláusula {num}: {titulo}", sClaus))
        story.append(Paragraph(texto, sNormal))

        # Insertar tabla de tarifas después de cláusula 11
        if num == "11":
            tarifas = _tarifas_driver(driver)
            if tarifas:
                story.append(Spacer(1, 6))
                tabla_data = [["Concepto", "Tarifa (CLP)"]]
                for concepto, valor in tarifas.items():
                    tabla_data.append([concepto, fmtCLP(valor)])
                t = Table(tabla_data, colWidths=[11*cm, 4*cm])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), AZUL),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), FONTB),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("FONTNAME", (0, 1), (-1, -1), FONT),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CLARO]),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ]))
                story.append(t)
                story.append(Spacer(1, 4))

    # Bloque de firma
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=1, color=AZUL))
    story.append(Spacer(1, 8))
    story.append(Paragraph("REGISTRO DE ACEPTACIÓN DIGITAL", ParagraphStyle("fh", fontName=FONTB, fontSize=10, leading=14, textColor=AZUL, alignment=TA_CENTER)))
    story.append(Spacer(1, 8))

    fecha_str = driver.acuerdo_fecha.strftime("%d-%m-%Y %H:%M") if driver.acuerdo_fecha else "—"
    meta_rows = [
        ["Prestador", driver.nombre or "—"],
        ["RUT", driver.rut or "—"],
        ["Versión aceptada", f"v{driver.acuerdo_version}" if driver.acuerdo_version else "—"],
        ["Fecha y hora de firma", fecha_str],
        ["Dirección IP", driver.acuerdo_ip or "—"],
    ]
    meta_table = Table(meta_rows, colWidths=[5*cm, 10*cm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), FONTB),
        ("FONTNAME", (1, 0), (1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TEXTCOLOR", (0, 0), (0, -1), AZUL),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.HexColor("#e5e7eb")),
    ]))
    story.append(meta_table)

    # Imagen de firma
    if driver.acuerdo_firma:
        try:
            raw = driver.acuerdo_firma
            if "," in raw:
                raw = raw.split(",", 1)[1]
            img_bytes = base64.b64decode(raw)
            img_buf = io.BytesIO(img_bytes)
            firma_img = RLImage(img_buf, width=6*cm, height=2.5*cm)
            story.append(Spacer(1, 10))
            story.append(Paragraph("Firma digital registrada:", sMeta))
            story.append(Spacer(1, 4))
            firma_wrap = Table([[firma_img]], colWidths=[7*cm])
            firma_wrap.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(firma_wrap)
        except Exception:
            pass

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "El presente registro de aceptación tiene validez legal conforme a la Ley N° 19.799 sobre Documentos "
        "Electrónicos, Firma Electrónica y Servicios de Certificación de la República de Chile.",
        sLegal,
    ))

    doc.build(story)
    return buf.getvalue()
def mi_acuerdo_tarifas(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Anexo de tarifas dinámico — siempre refleja la configuración actual."""
    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")
    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    return _tarifas_driver(driver)


def _tarifas_driver(driver: Driver) -> dict:
    tarifas = {}
    if driver.tarifa_ecourier:
        tarifas["Entrega estándar (Ecourier)"] = driver.tarifa_ecourier
    if driver.tarifa_oviedo:
        tarifas["Entrega Oviedo"] = driver.tarifa_oviedo
    if driver.tarifa_tercerizado:
        tarifas["Entrega tercerizado"] = driver.tarifa_tercerizado
    if driver.tarifa_valparaiso:
        tarifas["Entrega Valparaíso"] = driver.tarifa_valparaiso
    if driver.tarifa_melipilla:
        tarifas["Entrega Melipilla"] = driver.tarifa_melipilla
    if driver.tarifa_retiro_fija:
        tarifas["Retiro fijo"] = driver.tarifa_retiro_fija
    return tarifas


@router.get("/{driver_id}/acuerdo")
def acuerdo_del_driver(
    driver_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_administracion),
):
    """(Admin) Devuelve los detalles del acuerdo firmado por un driver."""
    from app.api.auth import CURRENT_ACUERDO_VERSION

    driver = db.get(Driver, driver_id)
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    return {
        "nombre": driver.nombre,
        "nombre_completo": driver.nombre_completo,
        "rut": driver.rut,
        "contratado": driver.contratado,
        "acuerdo_aceptado": bool(driver.contratado) or bool(driver.acuerdo_aceptado and driver.acuerdo_version == CURRENT_ACUERDO_VERSION),
        "acuerdo_version": driver.acuerdo_version,
        "acuerdo_fecha": driver.acuerdo_fecha.isoformat() if driver.acuerdo_fecha else None,
        "acuerdo_ip": driver.acuerdo_ip,
        "acuerdo_firma": driver.acuerdo_firma,
        "carnet_frontal": driver.carnet_frontal,
        "carnet_trasero": driver.carnet_trasero,
        "version_actual": CURRENT_ACUERDO_VERSION,
        "tarifas": _tarifas_driver(driver),
    }


@router.get("/me/contrato-trabajo-info")
def mi_contrato_trabajo_info(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Returns contract data for a contratado driver, pulling from linked Trabajador."""
    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")
    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    if not driver.contratado or not driver.trabajador_id:
        raise HTTPException(status_code=400, detail="Solo para conductores contratados vinculados")

    from app.models import Trabajador, ContratoTrabajadorVersion
    trabajador = db.get(Trabajador, driver.trabajador_id)
    if not trabajador:
        raise HTTPException(status_code=404, detail="Trabajador vinculado no encontrado")

    # Versión vigente del contrato (para jornada y distribución reales)
    version_vigente = (
        db.query(ContratoTrabajadorVersion)
        .filter(ContratoTrabajadorVersion.trabajador_id == trabajador.id)
        .order_by(ContratoTrabajadorVersion.vigente_desde.desc())
        .first()
    )

    return {
        "driver_nombre": driver.nombre,
        "contrato_trabajo_aceptado": driver.contrato_trabajo_aceptado,
        "contrato_trabajo_version": driver.contrato_trabajo_version,
        "contrato_trabajo_fecha": driver.contrato_trabajo_fecha.isoformat() if driver.contrato_trabajo_fecha else None,
        "version_actual": CURRENT_CONTRATO_TRABAJO_VERSION,
        "trabajador": {
            "nombre": trabajador.nombre,
            "rut": trabajador.rut,
            "cargo": trabajador.cargo,
            "sueldo_base": trabajador.sueldo_base,
            "gratificacion": trabajador.gratificacion,
            "sueldo_bruto": trabajador.sueldo_bruto,
            "sueldo_liquido": trabajador.sueldo_liquido,
            "movilizacion": trabajador.movilizacion,
            "colacion": trabajador.colacion,
            "viaticos": trabajador.viaticos,
            "afp": trabajador.afp,
            "sistema_salud": trabajador.sistema_salud,
            "monto_cotizacion_salud": trabajador.monto_cotizacion_salud,
            "fecha_ingreso": trabajador.fecha_ingreso.isoformat() if trabajador.fecha_ingreso else None,
            "tipo_contrato": trabajador.tipo_contrato,
            "banco": trabajador.banco,
            "tipo_cuenta": trabajador.tipo_cuenta,
            "numero_cuenta": trabajador.numero_cuenta,
            # Datos provenientes de la versión vigente del contrato
            "jornada_semanal_horas": version_vigente.jornada_semanal_horas if version_vigente else None,
            "distribucion_jornada": version_vigente.distribucion_jornada if version_vigente else None,
            "tipo_jornada": version_vigente.tipo_jornada if version_vigente else None,
        },
    }


@router.post("/me/contrato-trabajo")
def aceptar_contrato_trabajo(
    body: ContratoTrabajoAceptarRequest,
    request: Request,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Registers digital acceptance of employment contract for contratado drivers."""
    from datetime import datetime, timezone
    from app.auth import create_access_token
    from app.schemas import TokenResponse

    if user["rol"] != RolEnum.DRIVER:
        raise HTTPException(status_code=403, detail="Solo para drivers")

    driver = db.get(Driver, user["id"])
    if not driver:
        raise HTTPException(status_code=404, detail="Driver no encontrado")
    if not driver.contratado or not driver.trabajador_id:
        raise HTTPException(status_code=400, detail="Solo para conductores contratados vinculados")

    # Validar que exista un anexo CONTRATO_INICIAL en estado EMITIDO o INFORMATIVO
    # listo para firmar. Si está en BORRADOR, admin todavía no aprobó la emisión.
    from app.models import AnexoContrato, EstadoAnexoEnum, TipoAnexoEnum
    anexo_listo = (
        db.query(AnexoContrato)
        .filter(
            AnexoContrato.trabajador_id == driver.trabajador_id,
            AnexoContrato.tipo == TipoAnexoEnum.CONTRATO_INICIAL.value,
            AnexoContrato.estado.in_([EstadoAnexoEnum.EMITIDO.value, EstadoAnexoEnum.INFORMATIVO.value]),
        )
        .order_by(AnexoContrato.created_at.desc())
        .first()
    )
    if not anexo_listo:
        raise HTTPException(
            status_code=403,
            detail=(
                "Tu contrato aún no ha sido aprobado por administración. "
                "Espera la notificación cuando esté listo para firmar."
            ),
        )

    # Validar que nombre y RUT ingresados coincidan con los del trabajador
    # vinculado al contrato. Evita firmas "por otro" y errores de tipeo.
    from app.models import Trabajador
    trabajador = db.get(Trabajador, driver.trabajador_id)
    if not trabajador:
        raise HTTPException(status_code=404, detail="Trabajador vinculado no encontrado")

    rut_ingresado = _normalizar_rut(body.rut)
    rut_contrato = _normalizar_rut(trabajador.rut)
    if not rut_contrato:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tu contrato no tiene RUT registrado. Contacta a RR.HH. antes de firmar."
            ),
        )
    if rut_ingresado != rut_contrato:
        raise HTTPException(
            status_code=400,
            detail=(
                "El RUT ingresado no coincide con el registrado en tu contrato. "
                "Revisa que coincida con el que figura en tu cédula de identidad."
            ),
        )

    nombre_ingresado = _normalizar_nombre(body.nombre_completo)
    nombre_contrato = _normalizar_nombre(trabajador.nombre)
    if not nombre_contrato:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tu contrato no tiene nombre registrado. Contacta a RR.HH. antes de firmar."
            ),
        )
    if nombre_ingresado != nombre_contrato:
        raise HTTPException(
            status_code=400,
            detail=(
                "El nombre ingresado no coincide con el registrado en tu contrato. "
                f"Debe ser exactamente: «{trabajador.nombre}». "
                "Si hay un error en el nombre registrado, contacta a RR.HH."
            ),
        )

    ip = request.headers.get("X-Forwarded-For", "")
    if ip:
        ip = ip.split(",")[0].strip()
    else:
        ip = request.client.host if request.client else "unknown"

    driver.nombre_completo = body.nombre_completo
    driver.rut = body.rut
    driver.contrato_trabajo_aceptado = True
    driver.contrato_trabajo_version = CURRENT_CONTRATO_TRABAJO_VERSION
    driver.contrato_trabajo_fecha = datetime.now(timezone.utc)
    driver.contrato_trabajo_ip = ip
    driver.contrato_trabajo_firma = body.firma_base64
    driver.carnet_frontal = body.carnet_frontal
    driver.carnet_trasero = body.carnet_trasero

    # Sincronizar el anexo como FIRMADO para mantener una sola fuente de verdad
    anexo_listo.firma_trabajador_snapshot = body.firma_base64
    anexo_listo.firmado_at = datetime.now(timezone.utc)
    anexo_listo.firmado_ip = ip
    anexo_listo.estado = EstadoAnexoEnum.FIRMADO.value

    audit(
        db, "aceptar_contrato_trabajo",
        usuario=user, request=request,
        entidad="driver", entidad_id=driver.id,
        metadata={"version": CURRENT_CONTRATO_TRABAJO_VERSION, "ip": ip, "rut": body.rut},
    )
    db.commit()
    db.refresh(driver)

    token = create_access_token({"sub": str(driver.id), "rol": RolEnum.DRIVER})
    es_jefe = db.query(Driver).filter(Driver.jefe_flota_id == driver.id, Driver.activo == True).count() > 0
    return TokenResponse(
        access_token=token,
        rol=RolEnum.DRIVER,
        nombre=driver.nombre,
        entidad_id=driver.id,
        acuerdo_aceptado=True,
        contratado=driver.contratado,
        contrato_trabajo_aceptado=True,
        es_jefe=es_jefe,
    )
