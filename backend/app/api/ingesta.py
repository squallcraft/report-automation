from typing import List, Optional

import io
import os
import shutil
import uuid
import threading

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db, SessionLocal
from app.auth import require_admin, require_admin_or_administracion
from app.config import get_settings
from app.models import Envio, Seller, Driver, LogIngesta, RecepcionPaquete, Pickup
from app.schemas import IngestaResult, HomologacionPendiente, ResolverHomologacion
from app.services.audit import registrar as audit
from app.services.ingesta import procesar_reporte_excel, resolver_homologacion
from app.services.task_progress import create_task, get_task, update_task, cleanup_old_tasks

router = APIRouter(prefix="/ingesta", tags=["Ingesta"])


def _run_ingesta_in_background(
    filepath: str,
    task_id: str,
    usuario: str,
    usuario_dict: dict,
    reprocesar_semana: int = None,
    reprocesar_mes: int = None,
    reprocesar_anio: int = None,
):
    """Ejecuta la ingesta en un hilo separado con su propia sesión de BD."""
    db = SessionLocal()
    try:
        procesar_reporte_excel(
            db, filepath,
            usuario=usuario,
            task_id=task_id,
            reprocesar_semana=reprocesar_semana,
            reprocesar_mes=reprocesar_mes,
            reprocesar_anio=reprocesar_anio,
        )
        log = db.query(LogIngesta).filter(LogIngesta.ingesta_id == task_id).first()
        print(f"[INGESTA] Completada task={task_id}, registrando auditoría...", flush=True)
        audit(
            db, "ingesta_batch",
            usuario=usuario_dict,
            entidad="envio_batch",
            metadata={
                "archivo": os.path.basename(filepath),
                "task_id": task_id,
                "total_filas": log.total_filas if log else 0,
                "procesados": log.procesados if log else 0,
                "errores": log.errores_count if log else 0,
                "reprocesar": {"semana": reprocesar_semana, "mes": reprocesar_mes, "anio": reprocesar_anio}
                if reprocesar_semana else None,
            },
        )
        print(f"[INGESTA] Auditoría registrada para task={task_id}", flush=True)
    except Exception as e:
        import traceback
        print(f"[INGESTA ERROR] task={task_id}: {e}", flush=True)
        traceback.print_exc()
        update_task(task_id, status="error", message=f"Error fatal: {str(e)}")
    finally:
        db.close()


@router.get("/plantilla")
def descargar_plantilla():
    """Genera y descarga un archivo Excel plantilla con las columnas requeridas y datos de ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Envíos"

    columnas = [
        "User - Nombre",
        "Pedido Fecha",
        "Fecha Entrega",
        "Tracking ID",
        "Seller Code",
        "Seller Name",
        "External ID",
        "External Costo Orden",
        "Dirección",
        "Comuna",
        "Cantidad de Bultos",
        "Descripción Paquete",
        "Ruta Nombre",
        "Nombre Conductor",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col_idx, nombre in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ejemplo = [
        ["Carlos Pérez", "2026-02-24", "2026-02-25", "TRK-00001", "SC-10045", "MercadoLibre Chile", "ML-V-123456", 25990, "Av. Providencia 1234, Providencia, Santiago, Chile", "Providencia", 1, "[MLC1774402962] - Crema Reductora 500ml", "Ruta Santiago Centro", "Carlos Pérez"],
        ["Miguel López", "2026-02-24", "2026-02-25", "TRK-00002", "SC-10078", "Falabella", "FAL-789012", 45990, "Los Leones 567, Las Condes, Santiago, Chile", "Las Condes", 2, "Zapatillas deportivas talla 42", "Ruta Las Condes", "Miguel López"],
        ["Fernando Rojas", "2026-02-24", "2026-02-26", "TRK-00003", "SC-20012", "Ferretería Oviedo", "OV-345678", 12500, "Camino La Estrella 890, Padre Hurtado, Santiago, Chile", "Padre Hurtado", 1, "[MLC2220238846] - Taladro percutor 800W", "Ruta Sur", "Fernando Rojas"],
        ["Augusto Silva", "2026-02-25", "2026-02-26", "TRK-00004", "SC-30099", "Aventura Store", "AS-901234", 89990, "Av. Matta 2345, Santiago, Santiago, Chile", "Santiago", 3, "Carpa camping 4 personas", "Ruta Santiago Sur", "Augusto Silva"],
        ["Carlos Pérez", "2026-02-25", "2026-02-27", "TRK-00005", "SC-10046", "MercadoLibre Chile", "ML-V-567890", 15990, "Av. Macul 4567, La Florida, Santiago, Chile", "La Florida", 1, "Set de ollas antiadherentes", "Ruta Santiago Centro", "Carlos Pérez"],
    ]

    data_font = Font(size=10)
    data_align = Alignment(vertical="center")
    alt_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

    for row_idx, fila in enumerate(ejemplo, 2):
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    anchos = [22, 14, 14, 16, 14, 22, 18, 18, 50, 18, 16, 45, 22, 22]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho

    ws.auto_filter.ref = "A1:N1"
    ws.freeze_panes = "A2"

    ws_instrucciones = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("Instrucciones para el Reporte de Envíos ECourier", ""),
        ("", ""),
        ("Columna", "Descripción"),
        ("User - Nombre", "Nombre del driver que realizó la entrega. Debe coincidir con un nombre o alias registrado en el sistema."),
        ("Pedido Fecha", "Fecha en que se cargó el paquete. Formato: AAAA-MM-DD"),
        ("Fecha Entrega", "Fecha efectiva de entrega (OBLIGATORIA). Base para la liquidación. Formato: AAAA-MM-DD"),
        ("Tracking ID", "Identificador único del envío en el sistema de tracking."),
        ("Seller Code", "Código del seller para trazabilidad. No se usa en cálculos, pero es visible en el detalle de envíos."),
        ("Seller Name", "Nombre del seller/tienda. Debe coincidir con un nombre o alias registrado en el sistema."),
        ("External ID", "ID de la venta en la plataforma del seller (MercadoLibre, etc.)"),
        ("External Costo Orden", "Valor declarado del producto en CLP (solo número, sin $ ni puntos)."),
        ("Dirección", "Dirección completa de entrega."),
        ("Comuna", "Comuna de destino (OBLIGATORIA para cálculo de tarifa). Se usa para determinar el cobro al seller según su plan tarifario."),
        ("Cantidad de Bultos", "Número de bultos del envío (por defecto 1)."),
        ("Descripción Paquete", "Descripción del producto. Si contiene un código [MLCxxxxxxx], se detecta para aplicar extras."),
        ("Ruta Nombre", "Nombre de la ruta asignada al driver."),
        ("Nombre Conductor", "Nombre completo del conductor que realizó la entrega."),
        ("", ""),
        ("Notas importantes:", ""),
        ("", "• La columna 'Fecha Entrega' es OBLIGATORIA. Filas sin esta fecha serán ignoradas."),
        ("", "• Los nombres de sellers y drivers se homologan automáticamente por nombre exacto o alias."),
        ("", "• Si un nombre no se reconoce, el envío queda 'sin homologar' para revisión manual."),
        ("", "• Los códigos MLC en la descripción se extraen con formato [MLCxxxxxxx]."),
        ("", "• Las comunas se normalizan a minúsculas para buscar tarifas especiales."),
    ]

    title_font = Font(bold=True, size=14, color="1A365D")
    sub_font = Font(bold=True, size=11, color="2D3748")
    normal_font = Font(size=10, color="4A5568")

    for row_idx, (col_a, col_b) in enumerate(instrucciones, 1):
        cell_a = ws_instrucciones.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws_instrucciones.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            cell_a.font = title_font
        elif row_idx == 3:
            cell_a.font = sub_font
            cell_b.font = sub_font
        elif col_a and row_idx > 3 and row_idx <= 15:
            cell_a.font = Font(bold=True, size=10, color="2D3748")
            cell_b.font = normal_font
        elif "Notas" in col_a:
            cell_a.font = sub_font
        else:
            cell_a.font = normal_font
            cell_b.font = normal_font

    ws_instrucciones.column_dimensions["A"].width = 25
    ws_instrucciones.column_dimensions["B"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_reporte_ecourier.xlsx"},
    )


@router.post("/upload")
async def subir_reporte(
    file: UploadFile = File(...),
    reprocesar_semana: Optional[int] = Query(None),
    reprocesar_mes: Optional[int] = Query(None),
    reprocesar_anio: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """
    Sube un reporte Excel y lo procesa en segundo plano.
    Retorna un task_id para consultar el progreso via GET /ingesta/progress/{task_id}.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")

    safe_name = os.path.basename(file.filename)
    settings = get_settings()
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(settings.UPLOAD_DIR, safe_name)

    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    task_id = str(uuid.uuid4())[:12]
    usuario = current_user.get("nombre", current_user.get("username", "desconocido"))

    create_task(task_id, total=0, archivo=safe_name)

    thread = threading.Thread(
        target=_run_ingesta_in_background,
        args=(filepath, task_id, usuario, dict(current_user), reprocesar_semana, reprocesar_mes, reprocesar_anio),
        daemon=True,
    )
    thread.start()

    cleanup_old_tasks()

    return {"task_id": task_id, "message": "Procesamiento iniciado"}


@router.get("/progress/{task_id}")
def obtener_progreso(task_id: str):
    """Retorna el estado actual de un proceso de ingesta en background."""
    task = get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    return task


@router.get("/pendientes", response_model=List[HomologacionPendiente])
def listar_pendientes(db: Session = Depends(get_db), _=Depends(require_admin)):
    pendientes = []

    sellers_pendientes = (
        db.query(Envio.seller_nombre_raw, sqlfunc.count(Envio.id))
        .filter(Envio.seller_id.is_(None), Envio.seller_nombre_raw.isnot(None))
        .group_by(Envio.seller_nombre_raw)
        .all()
    )
    for nombre, cantidad in sellers_pendientes:
        pendientes.append(HomologacionPendiente(
            nombre_raw=nombre, tipo="SELLER", cantidad=cantidad
        ))

    drivers_pendientes = (
        db.query(Envio.driver_nombre_raw, sqlfunc.count(Envio.id))
        .filter(Envio.driver_id.is_(None), Envio.driver_nombre_raw.isnot(None))
        .group_by(Envio.driver_nombre_raw)
        .all()
    )
    for nombre, cantidad in drivers_pendientes:
        pendientes.append(HomologacionPendiente(
            nombre_raw=nombre, tipo="DRIVER", cantidad=cantidad
        ))

    pickups_pendientes = (
        db.query(RecepcionPaquete.pickup_nombre_raw, sqlfunc.count(RecepcionPaquete.id))
        .filter(RecepcionPaquete.pickup_id.is_(None), RecepcionPaquete.pickup_nombre_raw.isnot(None))
        .group_by(RecepcionPaquete.pickup_nombre_raw)
        .all()
    )
    for nombre, cantidad in pickups_pendientes:
        pendientes.append(HomologacionPendiente(
            nombre_raw=nombre, tipo="PICKUP", cantidad=cantidad
        ))

    return pendientes


@router.get("/logs")
def listar_logs(
    limit: int = 50,
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_administracion),
):
    logs = db.query(LogIngesta).order_by(LogIngesta.created_at.desc()).limit(limit).all()
    return [{
        "id": l.id,
        "ingesta_id": l.ingesta_id,
        "usuario": l.usuario,
        "tipo": l.tipo,
        "archivo": l.archivo,
        "total_filas": l.total_filas,
        "procesados": l.procesados,
        "errores_count": l.errores_count,
        "sin_homologar_sellers": l.sin_homologar_sellers or [],
        "sin_homologar_drivers": l.sin_homologar_drivers or [],
        "errores": l.errores or [],
        "resultado": l.resultado or {},
        "created_at": l.created_at.isoformat() if l.created_at else None,
    } for l in logs]


@router.post("/resolver")
def resolver(
    data: ResolverHomologacion,
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    try:
        actualizados = resolver_homologacion(db, data.nombre_raw, data.tipo, data.entidad_id)
        audit(
            db, "resolver_homologacion",
            usuario=current_user, request=request,
            entidad=data.tipo.lower(), entidad_id=data.entidad_id,
            metadata={"nombre_raw": data.nombre_raw, "actualizados": actualizados},
        )
        return {"message": f"{actualizados} envíos actualizados", "actualizados": actualizados}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/periodos-disponibles")
def periodos_disponibles(db: Session = Depends(get_db), _=Depends(require_admin_or_administracion)):
    """Retorna los períodos (semana, mes, año) que tienen envíos cargados."""
    rows = (
        db.query(Envio.semana, Envio.mes, Envio.anio, sqlfunc.count(Envio.id))
        .group_by(Envio.semana, Envio.mes, Envio.anio)
        .order_by(Envio.anio.desc(), Envio.mes.desc(), Envio.semana.desc())
        .all()
    )
    return [
        {"semana": r[0], "mes": r[1], "anio": r[2], "total_envios": r[3]}
        for r in rows
    ]
